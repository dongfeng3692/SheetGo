"""Excel reader: calamine for fast data reads, openpyxl for format-aware reads."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from python_calamine import CalamineWorkbook

from .models import (
    CellRange,
    FormulaInfo,
    InvalidCellRefError,
    SheetNotFoundError,
    col_number,
    parse_cell_ref,
)


class ExcelReader:
    """Excel file reader with dual engines.

    - python-calamine (Rust) for fast bulk data reads
    - openpyxl for formulas, styles, merged cells, dimensions
    """

    # -- calamine methods ------------------------------------------------

    @staticmethod
    def read_sheet_names(file_path: str | Path) -> list[str]:
        """Return all sheet names in the workbook."""
        wb = CalamineWorkbook.from_path(str(file_path))
        return wb.sheet_names

    @staticmethod
    def read_sheet_data(
        file_path: str | Path,
        sheet: str,
        range: CellRange | None = None,
    ) -> pd.DataFrame:
        """Read sheet data as a DataFrame (calamine, very fast).

        If *range* is provided, filter to that rectangular region.
        First row is used as column headers.
        """
        file_path = str(file_path)
        wb = CalamineWorkbook.from_path(file_path)
        names = wb.sheet_names
        if sheet not in names:
            raise SheetNotFoundError(
                f"Sheet {sheet!r} not found. Available: {names}"
            )
        data = wb.get_sheet_by_name(sheet).to_python()
        if not data:
            return pd.DataFrame()

        # First row as header
        headers = data[0]
        rows = data[1:]

        df = pd.DataFrame(rows, columns=headers)

        if range is not None:
            start_c = col_number(range.start_col) - 1  # 0-based
            end_c = col_number(range.end_col)           # exclusive
            start_r = range.start_row - 2               # -1 for header, -1 for 0-based
            end_r = range.end_row - 1                   # -1 for header
            # Clamp
            start_c = max(0, start_c)
            start_r = max(0, start_r)
            end_c = min(len(df.columns), end_c)
            end_r = min(len(df), end_r)
            df = df.iloc[start_r:end_r, start_c:end_c]

        return df

    @staticmethod
    def read_cell(file_path: str | Path, sheet: str, cell: str) -> Any:
        """Read a single cell value (calamine)."""
        col_letters, row_num = parse_cell_ref(cell)
        wb = CalamineWorkbook.from_path(str(file_path))
        names = wb.sheet_names
        if sheet not in names:
            raise SheetNotFoundError(
                f"Sheet {sheet!r} not found. Available: {names}"
            )
        data = wb.get_sheet_by_name(sheet).to_python()

        row_idx = row_num - 1  # 0-based
        col_idx = col_number(col_letters) - 1

        if row_idx < 0 or row_idx >= len(data):
            return None
        row = data[row_idx]
        if col_idx < 0 or col_idx >= len(row):
            return None
        return row[col_idx]

    @staticmethod
    def read_all_sheets(file_path: str | Path) -> dict[str, pd.DataFrame]:
        """Read all sheets into a dict of DataFrames."""
        file_path = str(file_path)
        wb = CalamineWorkbook.from_path(file_path)
        result: dict[str, pd.DataFrame] = {}
        for name in wb.sheet_names:
            data = wb.get_sheet_by_name(name).to_python()
            if not data:
                result[name] = pd.DataFrame()
                continue
            headers = data[0]
            rows = data[1:]
            result[name] = pd.DataFrame(rows, columns=headers)
        return result

    # -- openpyxl methods ------------------------------------------------

    @staticmethod
    def read_formulas(
        file_path: str | Path, sheet: str | None = None
    ) -> list[FormulaInfo]:
        """Read all formulas using openpyxl.

        If *sheet* is None, scan all sheets.
        Returns a list of FormulaInfo objects.
        """
        wb = load_workbook(str(file_path), data_only=False)
        sheets_to_scan = [sheet] if sheet else wb.sheetnames
        if sheet and sheet not in wb.sheetnames:
            wb.close()
            raise SheetNotFoundError(
                f"Sheet {sheet!r} not found. Available: {wb.sheetnames}"
            )

        formulas: list[FormulaInfo] = []
        for sname in sheets_to_scan:
            ws = wb[sname]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        # Detect shared formula
                        is_shared = False
                        shared_ref = None
                        # openpyxl stores shared formula info internally
                        if hasattr(cell, "value") and isinstance(cell.value, str):
                            pass  # formula text is in cell.value

                        formulas.append(
                            FormulaInfo(
                                sheet=sname,
                                cell=cell.coordinate,
                                formula=cell.value,
                                depends_on=[],
                                is_shared=is_shared,
                                shared_ref=shared_ref,
                            )
                        )
        wb.close()
        return formulas

    @staticmethod
    def read_merged_cells(
        file_path: str | Path, sheet: str
    ) -> list[str]:
        """Read merged cell ranges for a sheet. Returns ['A1:C3', ...]."""
        wb = load_workbook(str(file_path))
        if sheet not in wb.sheetnames:
            wb.close()
            raise SheetNotFoundError(
                f"Sheet {sheet!r} not found. Available: {wb.sheetnames}"
            )
        ws = wb[sheet]
        result = [str(r) for r in ws.merged_cells.ranges]
        wb.close()
        return result

    @staticmethod
    def read_styles(
        file_path: str | Path,
        sheet: str,
        range: CellRange | None = None,
    ) -> dict[str, dict]:
        """Read style info (font, fill, border, number_format) per cell.

        Returns {cell_ref: {font_name, font_size, bold, italic, font_color,
                           fill_color, number_format, ...}}.
        """
        wb = load_workbook(str(file_path))
        if sheet not in wb.sheetnames:
            wb.close()
            raise SheetNotFoundError(
                f"Sheet {sheet!r} not found. Available: {wb.sheetnames}"
            )
        ws = wb[sheet]

        # Determine iteration bounds
        if range is not None:
            min_col = col_number(range.start_col)
            max_col = col_number(range.end_col)
            min_row = range.start_row
            max_row = range.end_row
        else:
            min_col = ws.min_column or 1
            max_col = ws.max_column or 1
            min_row = ws.min_row or 1
            max_row = ws.max_row or 1

        result: dict[str, dict] = {}
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        ):
            for cell in row:
                info: dict[str, Any] = {}
                # Font
                if cell.font:
                    info["font_name"] = cell.font.name
                    info["font_size"] = cell.font.size
                    info["bold"] = cell.font.bold
                    info["italic"] = cell.font.italic
                    if cell.font.color and cell.font.color.rgb:
                        info["font_color"] = cell.font.color.rgb
                # Fill
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    info["fill_color"] = cell.fill.fgColor.rgb
                # Number format
                if cell.number_format:
                    info["number_format"] = cell.number_format
                # Horizontal/vertical alignment
                if cell.alignment:
                    info["horizontal"] = cell.alignment.horizontal
                    info["vertical"] = cell.alignment.vertical

                result[cell.coordinate] = info

        wb.close()
        return result

    @staticmethod
    def read_dimensions(file_path: str | Path, sheet: str) -> CellRange:
        """Read worksheet dimensions, returns CellRange."""
        try:
            wb = load_workbook(str(file_path))
            if sheet not in wb.sheetnames:
                wb.close()
                raise SheetNotFoundError(
                    f"Sheet {sheet!r} not found. Available: {wb.sheetnames}"
                )
            ws = wb[sheet]
            dims = ws.dimensions
            wb.close()
        except (AttributeError, TypeError):
            dims = None

        if not dims or ":" not in dims:
            # Fallback: use calamine to get actual bounds
            return _dimensions_from_calamine(str(file_path), sheet)

        left, right = dims.split(":", 1)
        start_col, start_row = parse_cell_ref(left)
        end_col, end_row = parse_cell_ref(right)

        return CellRange(
            sheet=sheet,
            start_col=start_col,
            start_row=start_row,
            end_col=end_col,
            end_row=end_row,
        )


def _dimensions_from_calamine(file_path: str, sheet: str) -> CellRange:
    """Fallback: compute dimensions from calamine data."""
    wb = CalamineWorkbook.from_path(file_path)
    if sheet not in wb.sheet_names:
        raise SheetNotFoundError(
            f"Sheet {sheet!r} not found. Available: {wb.sheet_names}"
        )
    data = wb.get_sheet_by_name(sheet).to_python()
    if not data:
        return CellRange(sheet=sheet, start_col="A", start_row=1,
                         end_col="A", end_row=1)
    from .models import col_letter
    max_col = col_letter(len(data[0])) if data[0] else "A"
    return CellRange(
        sheet=sheet,
        start_col="A",
        start_row=1,
        end_col=max_col,
        end_row=len(data),
    )
