"""Chart engine: create, list, and remove charts using openpyxl."""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.chart import (
    AreaChart,
    BarChart,
    LineChart,
    PieChart,
    Reference,
    ScatterChart,
)
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

from .models import CellRange, ChartConfig, ChartInfo, EditResult, SheetNotFoundError


class ChartEngine:
    """Chart creation and management using openpyxl."""

    # -- create chart ----------------------------------------------------

    @staticmethod
    def create_chart(
        file_path: str | Path, config: ChartConfig
    ) -> EditResult:
        """Create a chart and insert it into the worksheet."""
        file_path = str(file_path)
        backup = file_path + ".bak"
        shutil.copy2(file_path, backup)
        try:
            wb = load_workbook(file_path)
            if config.target_sheet not in wb.sheetnames:
                raise SheetNotFoundError(
                    f"Sheet {config.target_sheet!r} not found"
                )
            ws = wb[config.target_sheet]

            # Build chart object
            chart = _build_chart(config)

            # Set titles
            if config.title:
                chart.title = config.title
            if config.x_axis_title:
                chart.x_axis.title = config.x_axis_title
            if config.y_axis_title:
                chart.y_axis.title = config.y_axis_title

            # Data references
            src_range = config.source_range
            data_ref = Reference(
                ws,
                min_col=_col_to_num(src_range.start_col),
                min_row=src_range.start_row,
                max_col=_col_to_num(src_range.end_col),
                max_row=src_range.end_row,
            )

            # Categories = first column
            cats = Reference(
                ws,
                min_col=_col_to_num(src_range.start_col),
                min_row=src_range.start_row + 1,  # skip header
                max_row=src_range.end_row,
            )

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)

            if config.show_labels:
                for s in chart.series:
                    s.dLbls = DataLabelList()
                    s.dLbls.showVal = True

            # Size (cm to cm, openpyxl uses cm)
            chart.width = config.width
            chart.height = config.height

            # Insert
            ws.add_chart(chart, config.target_cell)
            wb.save(file_path)
            wb.close()

            return EditResult.ok(
                cells=[f"{config.target_sheet}!{config.target_cell}"],
                warnings=[],
            )
        except Exception:
            shutil.copy2(backup, file_path)
            raise
        finally:
            Path(backup).unlink(missing_ok=True)

    # -- list charts -----------------------------------------------------

    @staticmethod
    def list_charts(
        file_path: str | Path, sheet: str | None = None
    ) -> list[ChartInfo]:
        """List charts in the workbook."""
        wb = load_workbook(str(file_path))
        result: list[ChartInfo] = []
        sheets = [sheet] if sheet else wb.sheetnames

        for sname in sheets:
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]
            for chart in ws._charts:
                anchor = ""
                if hasattr(chart, 'anchor') and chart.anchor:
                    # Try to extract cell anchor
                    try:
                        anchor = str(chart.anchor)
                    except Exception:
                        anchor = ""

                chart_type = type(chart).__name__.replace("Chart", "").lower()
                title = chart.title if hasattr(chart, "title") else None

                result.append(ChartInfo(
                    sheet=sname,
                    anchor=anchor,
                    chart_type=chart_type,
                    title=title,
                ))

        wb.close()
        return result

    # -- remove chart ----------------------------------------------------

    @staticmethod
    def remove_chart(
        file_path: str | Path, sheet: str, chart_index: int
    ) -> EditResult:
        """Remove a chart by index from a sheet."""
        file_path = str(file_path)
        wb = load_workbook(file_path)
        if sheet not in wb.sheetnames:
            wb.close()
            raise SheetNotFoundError(f"Sheet {sheet!r} not found")

        ws = wb[sheet]
        charts = ws._charts
        if chart_index < 0 or chart_index >= len(charts):
            wb.close()
            return EditResult.fail(
                warnings=[f"Chart index {chart_index} out of range (0-{len(charts)-1})"]
            )

        charts.pop(chart_index)
        wb.save(file_path)
        wb.close()
        return EditResult.ok()


def _build_chart(config: ChartConfig):
    """Create an openpyxl chart object from config."""
    ct = config.chart_type.lower()
    if ct == "bar":
        return BarChart()
    elif ct == "line":
        return LineChart()
    elif ct == "pie":
        return PieChart()
    elif ct == "area":
        return AreaChart()
    elif ct == "scatter":
        return ScatterChart()
    else:
        # Default to bar
        return BarChart()


def _col_to_num(col: str) -> int:
    """Column letter to 1-based number."""
    n = 0
    for c in col.upper():
        n = n * 26 + (ord(c) - 64)
    return n
