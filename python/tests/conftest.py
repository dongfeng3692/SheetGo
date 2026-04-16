"""Shared test fixtures for the Excel engine tests."""

import os
import shutil
import tempfile

import pandas as pd
import pytest


@pytest.fixture
def tmp_dir():
    """Provide a temporary directory that is cleaned up after the test."""
    d = tempfile.mkdtemp(prefix="exceler_test_")
    yield d
    shutil.rmtree(d, ignore_errors=True)


@pytest.fixture
def simple_xlsx(tmp_dir):
    """Create a simple xlsx file with known data for testing."""
    from openpyxl import Workbook

    path = os.path.join(tmp_dir, "simple.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers
    ws["A1"] = "Name"
    ws["B1"] = "Amount"
    ws["C1"] = "Region"

    # Data rows
    data = [
        ("Alice", 100, "East"),
        ("Bob", 200, "West"),
        ("Charlie", 300, "East"),
        ("Diana", 400, "West"),
    ]
    for i, (name, amount, region) in enumerate(data, 2):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = amount
        ws[f"C{i}"] = region

    # Add a formula
    ws["B6"] = "=SUM(B2:B5)"

    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def unpacked_xlsx(simple_xlsx, tmp_dir):
    """Unpack simple.xlsx into a work directory."""
    from python.excel.xml_helpers import XMLHelpers

    work_dir = os.path.join(tmp_dir, "unpacked")
    XMLHelpers().unpack(simple_xlsx, work_dir)
    return work_dir


@pytest.fixture
def duckdb_with_data(tmp_dir):
    """Create a DuckDB file with test data."""
    import duckdb

    db_path = os.path.join(tmp_dir, "test.duckdb")
    con = duckdb.connect(db_path)
    con.execute("""
        CREATE TABLE sheet1 (
            name VARCHAR,
            amount INTEGER,
            region VARCHAR
        )
    """)
    con.execute("""
        INSERT INTO sheet1 VALUES
        ('Alice', 100, 'East'),
        ('Bob', 200, 'West'),
        ('Charlie', 300, 'East'),
        ('Diana', 400, 'West')
    """)
    con.close()
    return db_path


@pytest.fixture
def multi_sheet_xlsx(tmp_dir):
    """Create an xlsx with 3 sheets."""
    from openpyxl import Workbook

    path = os.path.join(tmp_dir, "multi_sheet.xlsx")
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Sales"
    ws1["A1"] = "Product"
    ws1["B1"] = "Revenue"
    ws1["A2"] = "Widget"
    ws1["B2"] = 1000
    ws1["A3"] = "Gadget"
    ws1["B3"] = 2000

    ws2 = wb.create_sheet("Expenses")
    ws2["A1"] = "Category"
    ws2["B1"] = "Amount"
    ws2["A2"] = "Rent"
    ws2["B2"] = 500
    ws2["A3"] = "Supplies"
    ws2["B3"] = 200

    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "Metric"
    ws3["B1"] = "Value"
    ws3["A2"] = "Total Revenue"
    ws3["B2"] = "=Sales!B2+Sales!B3"
    ws3["A3"] = "Total Expenses"
    ws3["B3"] = "=Expenses!B2+Expenses!B3"

    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def preload_config(tmp_dir, simple_xlsx):
    """Create a PreloadConfig pointing at the simple_xlsx fixture."""
    from python.preload.pipeline import PreloadConfig

    working = os.path.join(tmp_dir, "working")
    cache = os.path.join(tmp_dir, "cache")
    os.makedirs(working, exist_ok=True)
    os.makedirs(cache, exist_ok=True)
    file_id = "test_file_001"
    return PreloadConfig(
        file_id=file_id,
        source_path=simple_xlsx,
        working_path=os.path.join(working, f"{file_id}.xlsx"),
        duckdb_path=os.path.join(cache, f"{file_id}.duckdb"),
        schema_path=os.path.join(cache, f"{file_id}_schema.json"),
        stats_path=os.path.join(cache, f"{file_id}_stats.json"),
    )


@pytest.fixture
def multi_preload_config(tmp_dir, multi_sheet_xlsx):
    """Create a PreloadConfig for multi-sheet xlsx."""
    from python.preload.pipeline import PreloadConfig

    working = os.path.join(tmp_dir, "working_m")
    cache = os.path.join(tmp_dir, "cache_m")
    os.makedirs(working, exist_ok=True)
    os.makedirs(cache, exist_ok=True)
    file_id = "multi_001"
    return PreloadConfig(
        file_id=file_id,
        source_path=multi_sheet_xlsx,
        working_path=os.path.join(working, f"{file_id}.xlsx"),
        duckdb_path=os.path.join(cache, f"{file_id}.duckdb"),
        schema_path=os.path.join(cache, f"{file_id}_schema.json"),
        stats_path=os.path.join(cache, f"{file_id}_stats.json"),
    )
