"""Tests for the preload pipeline: schema, stats, formulas, styles, full pipeline."""

import json
import os

import pandas as pd
import pytest

from python.preload.schema_extractor import SchemaExtractor, SheetSchema
from python.preload.stats_calculator import StatsCalculator, DataQuality
from python.preload.formula_scanner import FormulaScanner
from python.preload.style_extractor import StyleExtractor
from python.preload.pipeline import PreloadPipeline, PreloadConfig, PreloadResult
from python.excel.duckdb_query import DuckDBQuery


# ===========================================================================
# StatsCalculator
# ===========================================================================

class TestStatsCalculator:
    def test_numeric_column_stats(self):
        series = pd.Series([10.0, 20.0, 30.0, 40.0, 50.0])
        stats = StatsCalculator.compute_column_stats(series)
        assert stats["min"] == 10.0
        assert stats["max"] == 50.0
        assert stats["mean"] == 30.0
        assert "std" in stats

    def test_categorical_column_stats(self):
        series = pd.Series(["East", "West", "East", "North", "East", "West"])
        stats = StatsCalculator.compute_column_stats(series)
        assert "topValues" in stats
        assert len(stats["topValues"]) <= 5
        assert stats["topValues"][0]["value"] == "East"
        assert stats["topValues"][0]["count"] == 3

    def test_empty_series(self):
        series = pd.Series(dtype=float)
        stats = StatsCalculator.compute_column_stats(series)
        assert "null_count" in stats

    def test_data_quality(self):
        data = {
            "Sheet1": pd.DataFrame({
                "A": [1, 2, None, 4],
                "B": ["x", "y", "z", "x"],
            }),
        }
        dq = StatsCalculator.compute_data_quality(data)
        assert dq.null_rate > 0
        assert isinstance(dq.duplicate_rows, int)

    def test_compute_file_stats(self):
        data = {
            "Sheet1": pd.DataFrame({"A": [1, 2, 3], "B": ["x", "y", "z"]}),
            "Sheet2": pd.DataFrame({"C": [10, 20]}),
        }
        stats = StatsCalculator.compute_file_stats("f1", data)
        assert stats.total_sheets == 2
        assert stats.total_rows == 5  # 3 + 2
        assert stats.total_cols == 3  # 2 + 1
        assert stats.total_formulas == 0

    def test_file_stats_to_dict(self):
        data = {"S": pd.DataFrame({"X": [1]})}
        stats = StatsCalculator.compute_file_stats("f1", data)
        d = stats.to_dict()
        assert d["fileId"] == "f1"
        assert "dataQuality" in d
        assert "formulaSummary" in d


# ===========================================================================
# SchemaExtractor
# ===========================================================================

class TestSchemaExtractor:
    def test_sanitize_table_name(self):
        assert SchemaExtractor._sanitize_table_name("Sheet1") == "sheet1"
        assert SchemaExtractor._sanitize_table_name("My Data 2024") == "my_data_2024"
        assert SchemaExtractor._sanitize_table_name("1Sheet") == "_1sheet"

    def test_extract_single_sheet(self, simple_xlsx):
        data = {"Sheet1": pd.DataFrame({
            "Name": ["Alice", "Bob"],
            "Amount": [100, 200],
        })}
        schemas = SchemaExtractor.extract(data)
        assert len(schemas) == 1
        assert schemas[0].name == "Sheet1"
        assert len(schemas[0].columns) == 2

    def test_col_letter_mapping(self):
        data = {"S": pd.DataFrame({"A": [1], "B": [2], "C": [3]})}
        schemas = SchemaExtractor.extract(data)
        cols = schemas[0].columns
        assert cols[0].col_letter == "A"
        assert cols[1].col_letter == "B"
        assert cols[2].col_letter == "C"

    def test_null_counts(self):
        data = {"S": pd.DataFrame({"X": [1, None, 3], "Y": ["a", "b", None]})}
        schemas = SchemaExtractor.extract(data)
        assert schemas[0].columns[0].null_count == 1
        assert schemas[0].columns[1].null_count == 1

    def test_unique_counts(self):
        data = {"S": pd.DataFrame({"X": [1, 1, 2, 3]})}
        schemas = SchemaExtractor.extract(data)
        assert schemas[0].columns[0].unique_count == 3

    def test_data_range(self):
        assert SchemaExtractor._compute_data_range(100, 3) == "A1:C101"
        assert SchemaExtractor._compute_data_range(0, 0) == "A1"

    def test_to_dict(self):
        data = {"S": pd.DataFrame({"X": [1]})}
        schemas = SchemaExtractor.extract(data)
        d = schemas[0].to_dict()
        assert d["name"] == "S"
        assert d["tableName"] == "s"
        assert len(d["columns"]) == 1

    def test_duplicate_headers(self):
        data = {
            "S": pd.DataFrame(
                [[1, 10], [None, 20]],
                columns=["Amount", "Amount"],
            )
        }
        schemas = SchemaExtractor.extract(data)
        cols = schemas[0].columns
        assert len(cols) == 2
        assert cols[0].name == "Amount"
        assert cols[1].name == "Amount"
        assert cols[0].null_count == 1
        assert cols[1].null_count == 0


# ===========================================================================
# FormulaScanner
# ===========================================================================

class TestFormulaScanner:
    def test_find_sum_formula(self, simple_xlsx):
        result = FormulaScanner.scan(simple_xlsx)
        assert result.total_count >= 1
        # Should find at least one formula on Sheet1
        assert len(result.sheets) >= 1
        formulas = result.sheets[0].formulas
        assert any("SUM" in f.formula for f in formulas)

    def test_dependency_graph(self, simple_xlsx):
        result = FormulaScanner.scan(simple_xlsx)
        assert len(result.dependency_graph) >= 1

    def test_cross_sheet_refs(self, multi_sheet_xlsx):
        result = FormulaScanner.scan(multi_sheet_xlsx)
        assert result.total_count >= 2
        assert result.cross_sheet_count >= 2  # Summary refs Sales + Expenses

    def test_no_formulas(self, tmp_dir):
        """Plain data file should have no formulas."""
        from openpyxl import Workbook
        path = os.path.join(tmp_dir, "plain.xlsx")
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Hello"
        ws["B1"] = 42
        wb.save(path)
        wb.close()

        result = FormulaScanner.scan(path)
        assert result.total_count == 0


# ===========================================================================
# StyleExtractor
# ===========================================================================

class TestStyleExtractor:
    def test_extract_styles(self, simple_xlsx):
        result = StyleExtractor.extract(simple_xlsx)
        assert len(result.sheets) >= 1
        # Should have some style info for cells
        sheet = result.sheets[0]
        assert sheet.sheet == "Sheet1"

    def test_style_to_dict(self, simple_xlsx):
        from python.preload.style_extractor import CellStyle
        cs = CellStyle(font_name="Arial", bold=True, number_format="0.0%")
        d = cs.to_dict()
        assert d["font_name"] == "Arial"
        assert d["bold"] is True
        assert d["number_format"] == "0.0%"
        # Omit None values
        assert "fill_color" not in d


# ===========================================================================
# PreloadPipeline (end-to-end)
# ===========================================================================

class TestPreloadPipeline:
    def test_full_run(self, preload_config):
        pipeline = PreloadPipeline(preload_config)
        result = pipeline.run()

        assert result.status == "ok"
        assert result.duration_ms > 0
        assert os.path.isfile(preload_config.working_path)
        assert os.path.isfile(preload_config.duckdb_path)
        assert os.path.isfile(preload_config.schema_path)
        assert os.path.isfile(preload_config.stats_path)

    def test_progress_callback(self, preload_config):
        events: list[tuple] = []
        def on_progress(stage, pct, msg, elapsed):
            events.append((stage, pct, msg, elapsed))

        pipeline = PreloadPipeline(preload_config)
        result = pipeline.run(on_progress=on_progress)

        assert result.status == "ok"
        assert len(events) >= 5  # At least 5 progress events
        # Should see key stages
        stages = [e[0] for e in events]
        assert "reading" in stages
        assert "done" in stages

    def test_duckdb_tables(self, preload_config):
        pipeline = PreloadPipeline(preload_config)
        result = pipeline.run()
        assert result.status == "ok"

        tables = DuckDBQuery.list_tables(preload_config.duckdb_path)
        assert "sheet1" in tables

    def test_schema_json_structure(self, preload_config):
        pipeline = PreloadPipeline(preload_config)
        result = pipeline.run()
        assert result.status == "ok"

        schema = PreloadPipeline.get_schema(preload_config.schema_path)
        assert schema is not None
        assert schema["fileId"] == preload_config.file_id
        assert len(schema["sheets"]) >= 1
        sheet = schema["sheets"][0]
        assert sheet["name"] == "Sheet1"
        assert "columns" in sheet
        assert len(sheet["columns"]) >= 2
        # Check column has required fields
        col = sheet["columns"][0]
        assert "name" in col
        assert "colLetter" in col
        assert "dtype" in col
        assert "nullCount" in col

    def test_stats_json_structure(self, preload_config):
        pipeline = PreloadPipeline(preload_config)
        result = pipeline.run()
        assert result.status == "ok"

        stats = PreloadPipeline.get_stats(preload_config.stats_path)
        assert stats is not None
        assert stats["fileId"] == preload_config.file_id
        assert "totalSheets" in stats
        assert "dataQuality" in stats
        assert "formulaSummary" in stats

    def test_cache_read_none_when_missing(self, tmp_dir):
        assert PreloadPipeline.get_schema("/nonexistent.json") is None
        assert PreloadPipeline.get_stats("/nonexistent.json") is None

    def test_idempotent_rerun(self, preload_config):
        pipeline = PreloadPipeline(preload_config)
        r1 = pipeline.run()
        assert r1.status == "ok"

        # Run again — should overwrite without error
        pipeline2 = PreloadPipeline(preload_config)
        r2 = pipeline2.run()
        assert r2.status == "ok"

    def test_multi_sheet(self, multi_preload_config):
        pipeline = PreloadPipeline(multi_preload_config)
        result = pipeline.run()
        assert result.status == "ok"

        tables = DuckDBQuery.list_tables(multi_preload_config.duckdb_path)
        assert "sales" in tables
        assert "expenses" in tables
        assert "summary" in tables

        schema = PreloadPipeline.get_schema(multi_preload_config.schema_path)
        assert len(schema["sheets"]) == 3

    def test_error_handling(self, tmp_dir):
        cfg = PreloadConfig(
            file_id="bad",
            source_path="/nonexistent_file.xlsx",
            working_path=os.path.join(tmp_dir, "working", "bad.xlsx"),
            duckdb_path=os.path.join(tmp_dir, "cache", "bad.duckdb"),
            schema_path=os.path.join(tmp_dir, "cache", "bad_schema.json"),
            stats_path=os.path.join(tmp_dir, "cache", "bad_stats.json"),
        )
        pipeline = PreloadPipeline(cfg)
        result = pipeline.run()
        assert result.status == "error"
        assert result.error_message is not None

    def test_duplicate_headers_workbook(self, tmp_dir):
        from openpyxl import Workbook

        source = os.path.join(tmp_dir, "duplicate_headers.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Amount"
        ws["B1"] = "Amount"
        ws["A2"] = 100
        ws["B2"] = 200
        ws["A3"] = None
        ws["B3"] = 300
        wb.save(source)
        wb.close()

        cache_dir = os.path.join(tmp_dir, "cache_dup")
        working_dir = os.path.join(tmp_dir, "working_dup")
        os.makedirs(cache_dir, exist_ok=True)
        os.makedirs(working_dir, exist_ok=True)

        cfg = PreloadConfig(
            file_id="dup_headers",
            source_path=source,
            working_path=os.path.join(working_dir, "dup_headers.xlsx"),
            duckdb_path=os.path.join(cache_dir, "dup_headers.duckdb"),
            schema_path=os.path.join(cache_dir, "dup_headers_schema.json"),
            stats_path=os.path.join(cache_dir, "dup_headers_stats.json"),
        )

        result = PreloadPipeline(cfg).run()
        assert result.status == "ok"

        schema = PreloadPipeline.get_schema(cfg.schema_path)
        assert schema is not None
        columns = schema["sheets"][0]["columns"]
        assert [column["name"] for column in columns] == ["Amount", "Amount"]
