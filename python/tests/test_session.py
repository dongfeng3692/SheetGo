"""Session Store 测试"""

from __future__ import annotations

import os
import sys
import shutil

import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from session import (
    Database,
    FileRecord,
    MemoryManager,
    RollbackEngine,
    RollbackResult,
    Session,
    SnapshotManager,
)
from session.models import MessageRecord, SnapshotRecord, new_id, now_iso


# ============================================================================
# 辅助：创建测试 xlsx 文件
# ============================================================================


def _create_xlsx(path: str, data: dict[str, str] | None = None):
    """创建简单 xlsx 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    if data:
        for i, (k, v) in enumerate(data.items(), start=2):
            ws[f"A{i}"] = k
            ws[f"B{i}"] = v
    else:
        ws["A2"] = "test"
        ws["B2"] = 100
    wb.save(path)
    wb.close()


# ============================================================================
# Test: Database — Session CRUD
# ============================================================================


class TestSessionCRUD:
    def test_create_session(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session("Test Session")
        assert session.id
        assert session.name == "Test Session"
        assert session.created_at
        db.close()

    def test_get_session(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session("Test")
        fetched = db.get_session(session.id)
        assert fetched is not None
        assert fetched.name == "Test"
        db.close()

    def test_list_sessions(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        db.create_session("S1")
        db.create_session("S2")
        db.create_session("S3")
        sessions = db.list_sessions()
        assert len(sessions) == 3
        db.close()

    def test_update_session(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session("Old Name")
        db.update_session(session.id, name="New Name", total_tokens=100)
        updated = db.get_session(session.id)
        assert updated.name == "New Name"
        assert updated.total_tokens == 100
        db.close()

    def test_delete_session(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session("To Delete")
        db.delete_session(session.id)
        assert db.get_session(session.id) is None
        db.close()

    def test_delete_session_cascades(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session("Cascade")
        msg = MessageRecord(id=new_id(), session_id=session.id, role="user", content="hi")
        db.save_message(msg)
        db.save_memory(session.id, "preference", "lang", "zh")
        db.delete_session(session.id)
        assert db.get_messages(session.id) == []
        assert db.list_memories(session.id) == []
        db.close()


# ============================================================================
# Test: Database — Messages
# ============================================================================


class TestMessages:
    def test_save_and_get_messages(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()
        for i in range(5):
            db.save_message(MessageRecord(
                id=new_id(), session_id=session.id,
                role="user", content=f"msg_{i}",
            ))
        messages = db.get_messages(session.id)
        assert len(messages) == 5
        assert messages[0].content == "msg_0"
        assert messages[4].content == "msg_4"
        db.close()

    def test_get_recent_messages(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()
        for i in range(10):
            db.save_message(MessageRecord(
                id=new_id(), session_id=session.id,
                role="user", content=f"msg_{i}",
            ))
        recent = db.get_recent_messages(session.id, count=3)
        assert len(recent) == 3
        assert recent[-1].content == "msg_9"
        db.close()

    def test_delete_old_messages(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()
        for i in range(20):
            db.save_message(MessageRecord(
                id=new_id(), session_id=session.id,
                role="user", content=f"msg_{i}",
            ))
        deleted = db.delete_old_messages(session.id, keep_recent=5)
        assert deleted > 0
        remaining = db.get_messages(session.id)
        assert len(remaining) == 5
        db.close()

    def test_message_with_tool_calls(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()
        db.save_message(MessageRecord(
            id=new_id(), session_id=session.id,
            role="assistant", content="",
            tool_calls=[{"name": "query_data", "args": {"sql": "SELECT *"}}],
            tool_results=[{"result": "ok"}],
        ))
        messages = db.get_messages(session.id)
        assert messages[0].tool_calls is not None
        assert messages[0].tool_calls[0]["name"] == "query_data"
        assert messages[0].tool_results[0]["result"] == "ok"
        db.close()


# ============================================================================
# Test: Database — Memories
# ============================================================================


class TestMemories:
    def test_save_and_get_memory(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()
        db.save_memory(session.id, "preference", "lang", "zh")
        value = db.get_memory(session.id, "preference", "lang")
        assert value == "zh"
        db.close()

    def test_upsert_memory(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()
        db.save_memory(session.id, "preference", "lang", "zh")
        db.save_memory(session.id, "preference", "lang", "en")
        value = db.get_memory(session.id, "preference", "lang")
        assert value == "en"
        entries = db.list_memories(session.id, "preference")
        assert len(entries) == 1  # upsert 不创建重复
        db.close()

    def test_list_memories_by_category(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()
        db.save_memory(session.id, "preference", "a", 1)
        db.save_memory(session.id, "preference", "b", 2)
        db.save_memory(session.id, "pattern", "x", 3)
        prefs = db.list_memories(session.id, "preference")
        assert len(prefs) == 2
        all_mems = db.list_memories(session.id)
        assert len(all_mems) == 3
        db.close()


# ============================================================================
# Test: Database — File Records
# ============================================================================


class TestFileRecords:
    def test_save_and_get_file_record(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()
        record = FileRecord(
            file_id="f001", session_id=session.id,
            file_name="test.xlsx", file_size=1024,
            file_hash="abc123", file_type="xlsx",
            source_path="/source/test.xlsx",
            working_path="/working/test.xlsx",
        )
        db.save_file_record(record)
        fetched = db.get_file_record("f001")
        assert fetched is not None
        assert fetched.file_name == "test.xlsx"
        db.close()

    def test_list_file_records(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()
        for i in range(3):
            db.save_file_record(FileRecord(
                file_id=f"f{i}", session_id=session.id,
                file_name=f"file{i}.xlsx", file_size=100,
                file_hash=f"hash{i}", file_type="xlsx",
                source_path=f"/s/{i}", working_path=f"/w/{i}",
            ))
        records = db.list_file_records(session.id)
        assert len(records) == 3
        db.close()

    def test_update_preload_status(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()
        db.save_file_record(FileRecord(
            file_id="f001", session_id=session.id,
            file_name="test.xlsx", file_size=100,
            file_hash="h", file_type="xlsx",
            source_path="/s", working_path="/w",
        ))
        db.update_preload_status("f001", "done")
        fetched = db.get_file_record("f001")
        assert fetched.preload_status == "done"
        db.close()


# ============================================================================
# Test: MemoryManager
# ============================================================================


class TestMemoryManager:
    def test_get_conversation_context(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()
        for i in range(5):
            db.save_message(MessageRecord(
                id=new_id(), session_id=session.id,
                role="user" if i % 2 == 0 else "assistant",
                content=f"msg_{i}",
            ))
        mgr = MemoryManager(db)
        ctx = mgr.get_conversation_context(session.id, max_messages=3)
        assert len(ctx) == 3
        db.close()

    def test_remember_preference(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()
        mgr = MemoryManager(db)
        mgr.remember_preference(session.id, "language", "zh")
        prefs = mgr.get_preferences(session.id)
        assert prefs["language"] == "zh"
        db.close()

    def test_get_working_state(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()
        db.save_file_record(FileRecord(
            file_id="f1", session_id=session.id,
            file_name="a.xlsx", file_size=100,
            file_hash="h", file_type="xlsx",
            source_path="/s", working_path="/w",
            preload_status="done",
        ))
        mgr = MemoryManager(db)
        state = mgr.get_working_state(session.id)
        assert len(state["files"]) == 1
        assert state["files"][0]["preload_status"] == "done"
        db.close()


# ============================================================================
# Test: SnapshotManager + RollbackEngine
# ============================================================================


class TestSnapshotAndRollback:
    def test_create_snapshot(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()

        source = str(tmp_path / "source.xlsx")
        working = str(tmp_path / "working.xlsx")
        _create_xlsx(source)
        shutil.copy2(source, working)

        snap_mgr = SnapshotManager(db, str(tmp_path / "workspace"))
        snap = snap_mgr.create_snapshot(
            session_id=session.id,
            file_id="f001",
            description="初始上传",
            source_path=source,
            working_path=working,
        )
        assert snap.id
        assert snap.file_hash
        assert snap.description == "初始上传"
        db.close()

    def test_snapshot_chain(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()

        source = str(tmp_path / "source.xlsx")
        _create_xlsx(source, {"A": "1"})

        snap_mgr = SnapshotManager(db, str(tmp_path / "workspace"))

        # 创建 3 个快照
        for i in range(3):
            working = str(tmp_path / f"working_{i}.xlsx")
            shutil.copy2(source, working)
            # 修改 working
            wb = openpyxl.load_workbook(working)
            ws = wb.active
            ws[f"C{i+1}"] = f"step_{i}"
            wb.save(working)
            wb.close()

            snap_mgr.create_snapshot(
                session_id=session.id,
                file_id="f001",
                description=f"操作 {i+1}",
                source_path=source,
                working_path=working,
            )

        chain = snap_mgr.list_snapshots(session.id, "f001")
        assert len(chain) == 3
        assert chain[0].description == "操作 1"
        assert chain[2].description == "操作 3"

        # 验证 parent 链
        assert chain[1].parent_id == chain[0].id
        assert chain[2].parent_id == chain[1].id
        db.close()

    def test_rollback_to_snapshot(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()

        source = str(tmp_path / "source.xlsx")
        _create_xlsx(source, {"A": "1"})

        snap_mgr = SnapshotManager(db, str(tmp_path / "workspace"))
        working_path = str(tmp_path / "workspace" / session.id / "working" / "f001.xlsx")
        os.makedirs(os.path.dirname(working_path), exist_ok=True)

        # 创建 3 个快照
        snapshots = []
        for i in range(3):
            shutil.copy2(source, working_path)
            wb = openpyxl.load_workbook(working_path)
            ws = wb.active
            ws[f"C{i+1}"] = f"step_{i}"
            wb.save(working_path)
            wb.close()

            snap = snap_mgr.create_snapshot(
                session_id=session.id,
                file_id="f001",
                description=f"操作 {i+1}",
                source_path=source,
                working_path=working_path,
            )
            snapshots.append(snap)

        # 再做一次修改，让 working 不同于 snapshot 2
        wb = openpyxl.load_workbook(working_path)
        ws = wb.active
        ws["Z99"] = "extra"
        wb.save(working_path)
        wb.close()

        # 回滚到 snapshot 2（index=1）
        rb = RollbackEngine(db, snap_mgr)
        result = rb.rollback_to(
            session_id=session.id,
            file_id="f001",
            snapshot_id=snapshots[1].id,
            working_path=working_path,
        )
        assert result.success is True
        assert result.changes_lost == 1  # snapshot 3 的修改
        assert "操作 2" in result.message

        # 验证 working 文件已恢复到 snapshot 2 的状态
        wb = openpyxl.load_workbook(working_path)
        ws = wb.active
        assert ws["C2"].value == "step_1"
        assert ws["Z99"].value is None  # 额外修改已撤销
        wb.close()
        db.close()

    def test_rollback_preview(self, tmp_path):
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()

        source = str(tmp_path / "source.xlsx")
        _create_xlsx(source)

        snap_mgr = SnapshotManager(db, str(tmp_path / "workspace"))
        working_path = str(tmp_path / "workspace" / session.id / "working" / "f001.xlsx")
        os.makedirs(os.path.dirname(working_path), exist_ok=True)

        snapshots = []
        for i in range(4):
            shutil.copy2(source, working_path)
            wb = openpyxl.load_workbook(working_path)
            wb.active[f"C{i+1}"] = f"step_{i}"
            wb.save(working_path)
            wb.close()

            snap = snap_mgr.create_snapshot(
                session_id=session.id,
                file_id="f001",
                description=f"操作 {i+1}",
                source_path=source,
                working_path=working_path,
            )
            snapshots.append(snap)

        rb = RollbackEngine(db, snap_mgr)
        preview = rb.get_rollback_preview(session.id, "f001", snapshots[1].id)
        assert preview is not None
        assert len(preview.changes_to_restore) == 2   # 操作 1, 2
        assert len(preview.changes_to_lose) == 2       # 操作 3, 4
        db.close()

    def test_rollback_then_new_snapshot(self, tmp_path):
        """回滚后继续操作，新快照应追加到链中"""
        db = Database(str(tmp_path / "test.db"))
        session = db.create_session()

        source = str(tmp_path / "source.xlsx")
        _create_xlsx(source)

        snap_mgr = SnapshotManager(db, str(tmp_path / "workspace"))
        working_path = str(tmp_path / "workspace" / session.id / "working" / "f001.xlsx")
        os.makedirs(os.path.dirname(working_path), exist_ok=True)

        # 创建 2 个快照
        snaps = []
        for i in range(2):
            shutil.copy2(source, working_path)
            wb = openpyxl.load_workbook(working_path)
            wb.active[f"C{i+1}"] = f"step_{i}"
            wb.save(working_path)
            wb.close()

            snap = snap_mgr.create_snapshot(
                session_id=session.id,
                file_id="f001",
                description=f"操作 {i+1}",
                source_path=source,
                working_path=working_path,
            )
            snaps.append(snap)

        # 回滚到 snapshot 1
        rb = RollbackEngine(db, snap_mgr)
        rb.rollback_to(session.id, "f001", snaps[0].id, working_path)

        # 继续操作并创建新快照
        wb = openpyxl.load_workbook(working_path)
        wb.active["D1"] = "new_step"
        wb.save(working_path)
        wb.close()

        new_snap = snap_mgr.create_snapshot(
            session_id=session.id,
            file_id="f001",
            description="回滚后新操作",
            source_path=source,
            working_path=working_path,
        )

        # 链中现在有 3 个快照
        chain = snap_mgr.list_snapshots(session.id, "f001")
        assert len(chain) == 3
        assert chain[-1].description == "回滚后新操作"
        db.close()


# ============================================================================
# Test: Snapshot diff
# ============================================================================


class TestSnapshotDiff:
    def test_diff_clean_file(self, tmp_path):
        from session.snapshot import compute_diff

        path1 = str(tmp_path / "a.xlsx")
        path2 = str(tmp_path / "b.xlsx")
        _create_xlsx(path1)
        shutil.copy2(path1, path2)

        diff = compute_diff(path1, path2)
        assert diff["added_files"] == []
        assert diff["deleted_files"] == []
        assert diff["modified_files"] == {}

    def test_diff_modified_file(self, tmp_path):
        from session.snapshot import compute_diff

        path1 = str(tmp_path / "a.xlsx")
        path2 = str(tmp_path / "b.xlsx")
        _create_xlsx(path1, {"A": "1"})
        _create_xlsx(path2, {"A": "2"})

        diff = compute_diff(path1, path2)
        assert len(diff["modified_files"]) > 0
