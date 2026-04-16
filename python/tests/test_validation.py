"""Validation Engine 测试"""

from __future__ import annotations

import os
import sys
import tempfile
import zipfile

import openpyxl
import pandas as pd
import pytest

# 确保 python 包可导入
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from validation import ValidationEngine
from validation.result import ErrorCode, ValidationError, ValidationResult


# ============================================================================
# 测试用 Excel 文件工厂
# ============================================================================


def _create_clean_xlsx(path: str):
    """创建干净的 xlsx 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 表头
    ws["A1"] = "Name"
    ws["B1"] = "Score"
    ws["C1"] = "Grade"

    # 数据（纯值，无公式，确保 pandas 可正确读取）
    for i in range(2, 12):
        ws[f"A{i}"] = f"Student{i-1}"
        ws[f"B{i}"] = 70 + i
        ws[f"C{i}"] = "A" if 70 + i > 90 else "B"

    wb.save(path)
    wb.close()


def _create_formula_error_xlsx(path: str):
    """创建包含各种公式错误的 xlsx 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 先写入公式
    ws["A1"] = "Name"
    ws["B1"] = "Score"

    # 构造 #REF! — 删除引用的行后公式会变成 #REF!
    ws["A2"] = "test"
    ws["B2"] = 10
    ws["C2"] = "=A2+B2"

    # 删除行导致 #REF!（通过直接写入错误值模拟）
    ws["D2"] = "=B3"  # 正常引用
    # 手动设置错误值
    ws["E2"] = "#REF!"
    ws["E2"].data_type = "e"

    ws["E3"] = "#DIV/0!"
    ws["E3"].data_type = "e"

    ws["E4"] = "#VALUE!"
    ws["E4"].data_type = "e"

    ws["E5"] = "#NAME?"
    ws["E5"].data_type = "e"

    ws["E6"] = "#NULL!"
    ws["E6"].data_type = "e"

    ws["E7"] = "#NUM!"
    ws["E7"].data_type = "e"

    ws["E8"] = "#N/A"
    ws["E8"].data_type = "e"

    wb.save(path)
    wb.close()


def _create_compat_xlsx(path: str):
    """创建包含禁用函数的 xlsx 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Data"
    for i in range(2, 6):
        ws[f"A{i}"] = i

    # 禁用函数
    ws["B1"] = "=FILTER(A2:A5, A2:A5>2)"
    ws["C1"] = "=XLOOKUP(3, A2:A5, A2:A5)"
    ws["D1"] = "=UNIQUE(A2:A5)"
    ws["E1"] = "=SUM(A2:A5)"  # 正常函数

    wb.save(path)
    wb.close()


def _create_reference_issue_xlsx(path: str):
    """创建包含引用问题的 xlsx 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 表头
    ws["A1"] = "Name"
    ws["B1"] = "Value"

    # 数据行 2-6
    for i in range(2, 7):
        ws[f"A{i}"] = f"Item{i}"
        ws[f"B{i}"] = i * 10

    # 引用范围远超实际数据 (max_row=6, 引用到 100)
    ws["C2"] = "=SUM(B1:B100)"

    # 聚合函数范围太小（仅 2 单元格）
    ws["D2"] = "=SUM(B2:B3)"

    wb.save(path)
    wb.close()


def _create_data_quality_xlsx(path: str):
    """创建数据质量差的 xlsx 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["C1"] = "Mixed"

    # 大量空值（>30%）
    for i in range(2, 12):
        ws[f"A{i}"] = f"Item{i}" if i % 3 != 0 else None
        ws[f"B{i}"] = i * 10

    # 混合类型列
    ws["C2"] = 100
    ws["C3"] = "text"
    ws["C4"] = 200
    ws["C5"] = "more text"
    ws["C6"] = 300

    # 重复行
    ws["A12"] = "Item2"
    ws["B12"] = 20
    ws["C12"] = 100
    ws["A13"] = "Item2"
    ws["B13"] = 20
    ws["C13"] = 100

    wb.save(path)
    wb.close()


def _create_broken_xlsx(path: str):
    """创建损坏的 xlsx 文件"""
    with open(path, "wb") as f:
        f.write(b"this is not a zip file")


def _create_missing_files_xlsx(path: str):
    """创建缺少必需文件的 xlsx（有效 ZIP 但不是合法 xlsx）"""
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr("dummy.txt", "hello")


def _create_inconsistent_formula_xlsx(path: str):
    """创建同列公式模式不一致的 xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["C1"] = "Result"

    # 同列大部分公式是 =B{row}*2，少数偏离
    for i in range(2, 11):
        ws[f"A{i}"] = f"Item{i}"
        ws[f"B{i}"] = i * 10
        ws[f"C{i}"] = f"=B{i}*2"

    # 偏离主流模式的公式
    ws["C5"] = "=B5+100"
    ws["C8"] = "=B5*3"

    wb.save(path)
    wb.close()


# ============================================================================
# 测试：公式错误检测
# ============================================================================


class TestFormulaCheck:
    def test_clean_file_no_errors(self, tmp_path):
        path = str(tmp_path / "clean.xlsx")
        _create_clean_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_formulas(path)

        formula_errors = [e for e in errors if e.severity == "error"]
        assert len(formula_errors) == 0

    def test_detects_all_error_types(self, tmp_path):
        path = str(tmp_path / "formula_errors.xlsx")
        _create_formula_error_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_formulas(path)

        error_codes = {e.code for e in errors if e.severity == "error"}
        assert ErrorCode.FORMULA_REF_ERROR in error_codes
        assert ErrorCode.FORMULA_DIV_ZERO in error_codes
        assert ErrorCode.FORMULA_VALUE_ERROR in error_codes
        assert ErrorCode.FORMULA_NAME_ERROR in error_codes
        assert ErrorCode.FORMULA_NULL_ERROR in error_codes
        assert ErrorCode.FORMULA_NUM_ERROR in error_codes
        assert ErrorCode.FORMULA_NA_ERROR in error_codes

    def test_error_has_correct_category(self, tmp_path):
        path = str(tmp_path / "formula_errors.xlsx")
        _create_formula_error_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_formulas(path)

        for e in errors:
            assert e.category == "formula"

    def test_error_has_sheet_and_cell(self, tmp_path):
        path = str(tmp_path / "formula_errors.xlsx")
        _create_formula_error_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_formulas(path)

        for e in errors:
            assert e.sheet != ""
            assert e.cell != ""


# ============================================================================
# 测试：引用范围校验
# ============================================================================


class TestReferenceCheck:
    def test_out_of_range_reference(self, tmp_path):
        path = str(tmp_path / "ref_issues.xlsx")
        _create_reference_issue_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_references(path)

        out_of_range = [e for e in errors if e.code == ErrorCode.REF_OUT_OF_RANGE]
        assert len(out_of_range) > 0

    def test_insufficient_range(self, tmp_path):
        path = str(tmp_path / "ref_issues.xlsx")
        _create_reference_issue_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_references(path)

        insufficient = [e for e in errors if e.code == ErrorCode.REF_INSUFFICIENT_RANGE]
        assert len(insufficient) > 0

    def test_inconsistent_pattern(self, tmp_path):
        path = str(tmp_path / "inconsistent.xlsx")
        _create_inconsistent_formula_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_references(path)

        inconsistent = [e for e in errors if e.code == ErrorCode.REF_INCONSISTENT_PATTERN]
        assert len(inconsistent) > 0

    def test_clean_file_no_warnings(self, tmp_path):
        path = str(tmp_path / "clean.xlsx")
        _create_clean_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_references(path)

        assert len(errors) == 0


# ============================================================================
# 测试：函数兼容性检查
# ============================================================================


class TestCompatCheck:
    def test_detects_forbidden_functions(self, tmp_path):
        path = str(tmp_path / "compat.xlsx")
        _create_compat_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_compatibility(path)

        func_names = {e.detail.get("function") for e in errors}
        assert "FILTER" in func_names
        assert "XLOOKUP" in func_names
        assert "UNIQUE" in func_names

    def test_normal_function_not_flagged(self, tmp_path):
        path = str(tmp_path / "compat.xlsx")
        _create_compat_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_compatibility(path)

        func_names = {e.detail.get("function") for e in errors}
        assert "SUM" not in func_names

    def test_error_has_alternative(self, tmp_path):
        path = str(tmp_path / "compat.xlsx")
        _create_compat_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_compatibility(path)

        for e in errors:
            assert "alternative" in e.detail
            assert e.detail["alternative"] != ""


# ============================================================================
# 测试：数据质量评估
# ============================================================================


class TestDataQuality:
    def test_high_null_rate(self, tmp_path):
        path = str(tmp_path / "quality.xlsx")
        _create_data_quality_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_data_quality(path)

        null_errors = [e for e in errors if e.code == ErrorCode.QUALITY_HIGH_NULL_RATE]
        assert len(null_errors) > 0

    def test_duplicate_rows(self, tmp_path):
        path = str(tmp_path / "quality.xlsx")
        _create_data_quality_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_data_quality(path)

        dup_errors = [e for e in errors if e.code == ErrorCode.QUALITY_DUPLICATE_ROWS]
        assert len(dup_errors) > 0

    def test_mixed_types(self, tmp_path):
        path = str(tmp_path / "quality.xlsx")
        _create_data_quality_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_data_quality(path)

        mixed = [e for e in errors if e.code == ErrorCode.QUALITY_MIXED_TYPES]
        assert len(mixed) > 0

    def test_clean_file_no_quality_issues(self, tmp_path):
        path = str(tmp_path / "clean.xlsx")
        _create_clean_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_data_quality(path)

        # 干净文件不应有 warning 级别的质量问题
        warnings = [e for e in errors if e.severity == "warning"]
        assert len(warnings) == 0


# ============================================================================
# 测试：OpenXML 结构验证
# ============================================================================


class TestStructureCheck:
    def test_invalid_zip(self, tmp_path):
        path = str(tmp_path / "broken.xlsx")
        _create_broken_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_structure(path)

        assert any(e.code == ErrorCode.STRUCTURE_INVALID_ZIP for e in errors)

    def test_missing_required_files(self, tmp_path):
        path = str(tmp_path / "missing.xlsx")
        _create_missing_files_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_structure(path)

        codes = {e.code for e in errors}
        assert ErrorCode.STRUCTURE_MISSING_FILE in codes

    def test_valid_xlsx_no_structure_errors(self, tmp_path):
        path = str(tmp_path / "clean.xlsx")
        _create_clean_xlsx(path)

        engine = ValidationEngine()
        errors = engine.check_structure(path)

        assert len(errors) == 0


# ============================================================================
# 测试：ValidationEngine 整体流程
# ============================================================================


class TestValidationEngine:
    def test_full_check_clean_file(self, tmp_path):
        path = str(tmp_path / "clean.xlsx")
        _create_clean_xlsx(path)

        engine = ValidationEngine()
        result = engine.full_check(path, file_id="test_001")

        assert isinstance(result, ValidationResult)
        assert result.file_id == "test_001"
        assert result.timestamp != ""
        assert "formula_errors" in result.summary
        assert result.passed is True

    def test_full_check_problematic_file(self, tmp_path):
        path = str(tmp_path / "problems.xlsx")
        _create_formula_error_xlsx(path)

        engine = ValidationEngine()
        result = engine.full_check(path, file_id="test_002")

        assert result.error_count > 0
        assert result.passed is False

    def test_quick_check_specific_sheets(self, tmp_path):
        path = str(tmp_path / "multi_sheet.xlsx")
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Clean"
        ws1["A1"] = "Name"
        ws1["A2"] = "Alice"

        ws2 = wb.create_sheet("Problems")
        ws2["A1"] = "=FILTER(B:B, B:B>0)"

        wb.save(path)
        wb.close()

        engine = ValidationEngine()
        # 只检查 Clean sheet
        result = engine.quick_check(path, file_id="test_003", changed_sheets=["Clean"])

        # Clean sheet 不应有 compat 错误
        compat_errors = [e for e in result.errors if e.category == "compat"]
        assert len(compat_errors) == 0

    def test_final_check_structure_and_formula(self, tmp_path):
        path = str(tmp_path / "clean.xlsx")
        _create_clean_xlsx(path)

        engine = ValidationEngine()
        result = engine.final_check(path, file_id="test_004")

        # final_check 只做 structure + formula + reference
        categories = {e.category for e in result.errors}
        assert "quality" not in categories
        assert "compat" not in categories

    def test_result_passed_property(self):
        """passed=True 当且仅当没有 error 级别的问题"""
        r1 = ValidationResult(
            file_id="x", timestamp="t",
            errors=[ValidationError("warning", "formula", "S", "A1", "CODE", "msg")],
        )
        assert r1.passed is True

        r2 = ValidationResult(
            file_id="x", timestamp="t",
            errors=[ValidationError("error", "formula", "S", "A1", "CODE", "msg")],
        )
        assert r2.passed is False

        r3 = ValidationResult(
            file_id="x", timestamp="t",
            errors=[],
        )
        assert r3.passed is True

    def test_error_count_warning_count(self):
        errors = [
            ValidationError("error", "formula", "S", "A1", "C1", "m"),
            ValidationError("warning", "formula", "S", "A2", "C2", "m"),
            ValidationError("warning", "reference", "S", "A3", "C3", "m"),
            ValidationError("info", "quality", "S", "A4", "C4", "m"),
        ]
        r = ValidationResult(file_id="x", timestamp="t", errors=errors)
        assert r.error_count == 1
        assert r.warning_count == 2
