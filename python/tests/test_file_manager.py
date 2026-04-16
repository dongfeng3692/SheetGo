"""Tests for the file manager: hash, format detection, import, list, remove, export."""

import os
import shutil

import pytest

from python.file_manager.manager import FileError, FileManager
from python.session.database import Database
from python.session.models import new_id


@pytest.fixture
def db(tmp_path):
    """Create a temporary database."""
    db_path = str(tmp_path / "test.db")
    return Database(db_path)


@pytest.fixture
def fm(db):
    return FileManager(db)


@pytest.fixture
def session_id(db):
    """Create a session in DB and return its id (needed for FK constraint)."""
    session = db.create_session()
    return session.id


@pytest.fixture
def sample_xlsx(tmp_path):
    """Create a minimal xlsx file for testing."""
    from openpyxl import Workbook

    path = str(tmp_path / "sample.xlsx")
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Amount"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def sample_csv(tmp_path):
    path = str(tmp_path / "data.csv")
    with open(path, "w") as f:
        f.write("Name,Amount\nAlice,100\nBob,200\n")
    return path


@pytest.fixture
def sample_tsv(tmp_path):
    path = str(tmp_path / "data.tsv")
    with open(path, "w") as f:
        f.write("Name\tAmount\nAlice\t100\n")
    return path


@pytest.fixture
def fake_xlsx(tmp_path):
    """File with .xlsx extension but NOT valid ZIP content."""
    path = str(tmp_path / "fake.xlsx")
    with open(path, "wb") as f:
        f.write(b"this is not a zip file")
    return path


# ===========================================================================
# compute_hash
# ===========================================================================

class TestComputeHash:
    def test_deterministic(self, sample_xlsx):
        h1 = FileManager.compute_hash(sample_xlsx)
        h2 = FileManager.compute_hash(sample_xlsx)
        assert h1 == h2
        assert len(h1) == 64  # SHA-256 hex

    def test_different_files(self, sample_xlsx, sample_csv):
        h1 = FileManager.compute_hash(sample_xlsx)
        h2 = FileManager.compute_hash(sample_csv)
        assert h1 != h2


# ===========================================================================
# detect_format
# ===========================================================================

class TestDetectFormat:
    def test_xlsx(self, sample_xlsx):
        assert FileManager.detect_format(sample_xlsx) == "xlsx"

    def test_csv(self, sample_csv):
        assert FileManager.detect_format(sample_csv) == "csv"

    def test_tsv(self, sample_tsv):
        assert FileManager.detect_format(sample_tsv) == "tsv"

    def test_fake_xlsx(self, fake_xlsx):
        assert FileManager.detect_format(fake_xlsx) == "unknown"

    def test_unknown_extension(self, tmp_path):
        path = str(tmp_path / "file.txt")
        with open(path, "w") as f:
            f.write("hello")
        assert FileManager.detect_format(path) == "unknown"


# ===========================================================================
# import_file
# ===========================================================================

class TestImportFile:
    def test_basic_import(self, fm, db, session_id, sample_xlsx, tmp_path):
        file_id = new_id()
        working = str(tmp_path / "working" / "sample.xlsx")
        os.makedirs(os.path.dirname(working), exist_ok=True)
        shutil.copy2(sample_xlsx, working)

        result = fm.import_file(
            session_id=session_id,
            file_id=file_id,
            file_name="sample.xlsx",
            source_path=sample_xlsx,
            working_path=working,
        )

        assert result.file_id == file_id
        assert result.file_name == "sample.xlsx"
        assert result.file_type == "xlsx"
        assert result.file_hash != ""
        assert result.file_size > 0
        assert result.preload_status == "pending"
        assert result.duplicate_of is None

        # Verify DB record
        record = db.get_file_record(file_id)
        assert record is not None
        assert record.file_name == "sample.xlsx"

    def test_import_csv(self, fm, session_id, sample_csv):
        file_id = new_id()
        result = fm.import_file(
            session_id=session_id,
            file_id=file_id,
            file_name="data.csv",
            source_path=sample_csv,
            working_path=sample_csv,
        )
        assert result.file_type == "csv"

    def test_import_unsupported_format(self, fm, tmp_path):
        path = str(tmp_path / "file.txt")
        with open(path, "w") as f:
            f.write("hello")

        with pytest.raises(FileError, match="Unsupported"):
            fm.import_file(
                session_id=new_id(),
                file_id=new_id(),
                file_name="file.txt",
                source_path=path,
                working_path=path,
            )

    def test_import_fake_xlsx(self, fm, fake_xlsx):
        with pytest.raises(FileError, match="Unsupported"):
            fm.import_file(
                session_id=new_id(),
                file_id=new_id(),
                file_name="fake.xlsx",
                source_path=fake_xlsx,
                working_path=fake_xlsx,
            )

    def test_import_nonexistent_file(self, fm):
        with pytest.raises(FileError, match="not found"):
            fm.import_file(
                session_id=new_id(),
                file_id=new_id(),
                file_name="nope.xlsx",
                source_path="/nonexistent.xlsx",
                working_path="/nonexistent.xlsx",
            )

    def test_duplicate_detection(self, fm, session_id, sample_xlsx, tmp_path):
        working1 = str(tmp_path / "working1.xlsx")
        working2 = str(tmp_path / "working2.xlsx")
        shutil.copy2(sample_xlsx, working1)
        shutil.copy2(sample_xlsx, working2)

        r1 = fm.import_file(
            session_id=session_id,
            file_id=new_id(),
            file_name="first.xlsx",
            source_path=sample_xlsx,
            working_path=working1,
        )
        assert r1.duplicate_of is None

        r2 = fm.import_file(
            session_id=session_id,
            file_id=new_id(),
            file_name="second.xlsx",
            source_path=sample_xlsx,
            working_path=working2,
        )
        assert r2.duplicate_of == r1.file_id
        assert r2.preload_status == "duplicate"

    def test_different_sessions_no_duplicate(self, fm, db, sample_xlsx, tmp_path):
        """Same file in different sessions should NOT be flagged as duplicate."""
        sid1 = db.create_session().id
        sid2 = db.create_session().id
        working1 = str(tmp_path / "w1.xlsx")
        working2 = str(tmp_path / "w2.xlsx")
        shutil.copy2(sample_xlsx, working1)
        shutil.copy2(sample_xlsx, working2)

        r1 = fm.import_file(
            session_id=sid1,
            file_id=new_id(),
            file_name="a.xlsx",
            source_path=sample_xlsx,
            working_path=working1,
        )
        r2 = fm.import_file(
            session_id=sid2,
            file_id=new_id(),
            file_name="a.xlsx",
            source_path=sample_xlsx,
            working_path=working2,
        )
        assert r1.duplicate_of is None
        assert r2.duplicate_of is None


# ===========================================================================
# list_files
# ===========================================================================

class TestListFiles:
    def test_empty(self, fm):
        assert fm.list_files("nonexistent_session") == []

    def test_returns_imported(self, fm, session_id, sample_xlsx, tmp_path):
        working = str(tmp_path / "working.xlsx")
        shutil.copy2(sample_xlsx, working)

        fm.import_file(
            session_id=session_id,
            file_id="f1",
            file_name="test.xlsx",
            source_path=sample_xlsx,
            working_path=working,
        )

        files = fm.list_files(session_id)
        assert len(files) == 1
        assert files[0].file_id == "f1"


# ===========================================================================
# get_file_info
# ===========================================================================

class TestGetFileInfo:
    def test_found(self, fm, session_id, sample_xlsx, tmp_path):
        working = str(tmp_path / "working.xlsx")
        shutil.copy2(sample_xlsx, working)

        fm.import_file(
            session_id=session_id,
            file_id="f1",
            file_name="test.xlsx",
            source_path=sample_xlsx,
            working_path=working,
        )

        info = fm.get_file_info("f1")
        assert info is not None
        assert info.file_id == "f1"

    def test_not_found(self, fm):
        assert fm.get_file_info("nonexistent") is None


# ===========================================================================
# remove_file
# ===========================================================================

class TestRemoveFile:
    def test_removes_from_db(self, fm, db, session_id, sample_xlsx, tmp_path):
        working = str(tmp_path / "working.xlsx")
        shutil.copy2(sample_xlsx, working)

        fm.import_file(
            session_id=session_id,
            file_id="f1",
            file_name="test.xlsx",
            source_path=sample_xlsx,
            working_path=working,
        )

        fm.remove_file("f1", session_id)
        assert db.get_file_record("f1") is None

    def test_removes_working_file(self, fm, session_id, sample_xlsx, tmp_path):
        working = str(tmp_path / "working.xlsx")
        shutil.copy2(sample_xlsx, working)

        fm.import_file(
            session_id=session_id,
            file_id="f1",
            file_name="test.xlsx",
            source_path=sample_xlsx,
            working_path=working,
        )

        fm.remove_file("f1", session_id)
        assert not os.path.isfile(working)

    def test_not_found_raises(self, fm):
        with pytest.raises(FileError, match="not found"):
            fm.remove_file("nonexistent", "session")


# ===========================================================================
# export_file
# ===========================================================================

class TestExportFile:
    def test_exports_working_copy(self, fm, session_id, sample_xlsx, tmp_path):
        working = str(tmp_path / "working.xlsx")
        shutil.copy2(sample_xlsx, working)

        fm.import_file(
            session_id=session_id,
            file_id="f1",
            file_name="test.xlsx",
            source_path=sample_xlsx,
            working_path=working,
        )

        dest = str(tmp_path / "export.xlsx")
        fm.export_file("f1", dest)
        assert os.path.isfile(dest)
        # Should be the same content as working copy
        assert FileManager.compute_hash(dest) == FileManager.compute_hash(working)

    def test_export_nonexistent(self, fm):
        with pytest.raises(FileError, match="not found"):
            fm.export_file("nonexistent", "/tmp/out.xlsx")


# ===========================================================================
# update_preload_status
# ===========================================================================

class TestUpdatePreloadStatus:
    def test_update(self, fm, db, session_id, sample_xlsx, tmp_path):
        working = str(tmp_path / "working.xlsx")
        shutil.copy2(sample_xlsx, working)

        fm.import_file(
            session_id=session_id,
            file_id="f1",
            file_name="test.xlsx",
            source_path=sample_xlsx,
            working_path=working,
        )

        fm.update_preload_status("f1", "ready")
        record = db.get_file_record("f1")
        assert record.preload_status == "ready"
