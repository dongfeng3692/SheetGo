"""引用范围校验 — 检测 4 种引用问题"""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from typing import TYPE_CHECKING

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from .result import ErrorCode, ValidationError, safe_formula_str

if TYPE_CHECKING:
    from openpyxl import Workbook

# 聚合函数列表
_AGGREGATE_FUNCTIONS = frozenset({
    "SUM", "AVERAGE", "COUNT", "COUNTA", "COUNTBLANK",
    "MAX", "MIN", "PRODUCT", "STDEV", "STDEVP",
    "VAR", "VARP", "MEDIAN", "MODE",
})

# 单元格引用正则：匹配 A1:B2, $A$1:$B$2, A:A, 1:1 等模式
_CELL_REF_RE = re.compile(
    r"(?:'[^']+?'!)?"                        # 可选的 sheet 名（含引号）
    r"(\$?[A-Z]{1,3})?"                       # 可选的列字母
    r"(\$?\d+)?"                              # 可选的行号
    r"(?::(\$?[A-Z]{1,3})?(\$?\d+)?)?"       # 可选的范围结束部分
    , re.IGNORECASE
)

# 范围引用正则（A1:B2 形式）
_RANGE_REF_RE = re.compile(
    r"(?:'([^']+?)'!)?"                       # sheet 名
    r"\$?([A-Z]{1,3})\$?(\d+)"               # 起始列行
    r":"
    r"\$?([A-Z]{1,3})\$?(\d+)"               # 结束列行
    , re.IGNORECASE
)

# 整列/整行引用正则（A:A, 1:1 形式）
_WHOLE_COL_ROW_RE = re.compile(
    r"(?:'([^']+?)'!)?"                       # sheet 名
    r"\$?([A-Z]{1,3})?:\$?([A-Z]{1,3})?"     # 列范围 (A:A, A:B)
    r"|"
    r"\$?(\d+)?:\$?(\d+)?"                   # 行范围 (1:1, 1:10)
    , re.IGNORECASE
)

# 单个单元格引用正则
_SINGLE_CELL_RE = re.compile(
    r"(?:'([^']+?)'!)?"                       # sheet 名
    r"\$?([A-Z]{1,3})\$?(\d+)"               # 列行
    r"(?!:)"                                  # 不后面跟冒号（排除范围）
    , re.IGNORECASE
)


@dataclass
class CellRef:
    """单元格范围引用"""
    sheet: str | None
    min_col: str
    min_row: int
    max_col: str | None = None
    max_row: int | None = None

    @property
    def row_count(self) -> int:
        if self.max_row is not None and self.min_row is not None:
            return self.max_row - self.min_row + 1
        return 1


def extract_cell_references(formula: str) -> list[CellRef]:
    """从公式中提取所有单元格范围引用"""
    if not formula:
        return []

    refs: list[CellRef] = []

    # 提取范围引用 A1:B2
    for m in _RANGE_REF_RE.finditer(formula):
        sheet = m.group(1)
        min_col = m.group(2).upper()
        min_row = int(m.group(3))
        max_col = m.group(4).upper()
        max_row = int(m.group(5))
        refs.append(CellRef(
            sheet=sheet, min_col=min_col, min_row=min_row,
            max_col=max_col, max_row=max_row,
        ))

    # 提取整列引用 A:A
    for m in _WHOLE_COL_ROW_RE.finditer(formula):
        sheet = m.group(1)
        col_start = (m.group(2) or "").upper()
        col_end = (m.group(3) or "").upper()
        row_start = m.group(4)
        row_end = m.group(5)

        if col_start and col_end:
            refs.append(CellRef(
                sheet=sheet, min_col=col_start, min_row=1,
                max_col=col_end, max_row=1048576,
            ))
        elif row_start and row_end:
            refs.append(CellRef(
                sheet=sheet, min_col="A", min_row=int(row_start),
                max_col="XFD", max_row=int(row_end),
            ))

    return refs


def is_aggregate_function(formula: str) -> bool:
    """判断公式最外层是否为聚合函数"""
    if not formula or not formula.startswith("="):
        return False
    stripped = formula[1:].strip()
    for func in _AGGREGATE_FUNCTIONS:
        if stripped.upper().startswith(func + "("):
            return True
    return False


def extract_function_name(formula: str) -> str | None:
    """提取公式最外层的函数名"""
    if not formula or not formula.startswith("="):
        return None
    stripped = formula[1:].strip()
    m = re.match(r'([A-Z]+)\s*\(', stripped, re.IGNORECASE)
    return m.group(1).upper() if m else None


def _has_header_row(ws) -> bool:
    """启发式判断第一行是否为表头（文本，后续行为数据）"""
    if ws.max_row is None or ws.max_row < 2:
        return False
    first_row_types = set()
    second_row_types = set()
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        first_row_types.add(cell.data_type)
    for cell in next(ws.iter_rows(min_row=2, max_row=2)):
        second_row_types.add(cell.data_type)
    # 如果第一行是文本（s），第二行不是，认为是表头
    if "s" in first_row_types and "s" not in second_row_types:
        return True
    return False


def _check_formula_consistency(wb: Workbook, sheets: list[str] | None = None) -> list[ValidationError]:
    """检查同列公式模式一致性"""
    errors: list[ValidationError] = []
    target_sheets = sheets if sheets else [ws.title for ws in wb.worksheets]

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # 按列收集公式
        col_formulas: dict[str, list[tuple[int, str, str]]] = defaultdict(list)
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    col_letter = cell.column_letter
                    col_formulas[col_letter].append((cell.row, cell.coordinate, safe_formula_str(cell.value)))

        for col_letter, formulas in col_formulas.items():
            if len(formulas) < 3:
                continue

            # 将公式按行号排序
            formulas.sort(key=lambda x: x[0])

            # 提取模式：将公式中的行号替换为占位符
            def normalize(f: str) -> str:
                return re.sub(r'\d+', 'N', f)

            patterns: dict[str, int] = defaultdict(int)
            for _, coord, formula in formulas:
                patterns[normalize(formula)] += 1

            if len(patterns) <= 1:
                continue

            # 找出主流模式
            main_pattern = max(patterns, key=patterns.get)
            main_count = patterns[main_pattern]

            # 少数派（偏离主流）视为不一致
            for row_num, coord, formula in formulas:
                norm = normalize(formula)
                if norm != main_pattern and patterns[norm] < main_count * 0.3:
                    errors.append(ValidationError(
                        severity="warning",
                        category="reference",
                        sheet=sheet_name,
                        cell=coord,
                        code=ErrorCode.REF_INCONSISTENT_PATTERN,
                        message=f"同列公式模式不一致: {coord}={formula}（主流模式: {main_pattern}）",
                        detail={
                            "formula": formula,
                            "main_pattern": main_pattern,
                        },
                    ))

    return errors


def check_references(
    file_path: str,
    sheets: list[str] | None = None,
) -> list[ValidationError]:
    """
    4 种引用问题检测:
    1. 引用范围远超实际数据行
    2. 公式包含表头行
    3. 聚合函数范围太小（≤2 单元格）
    4. 同列公式模式不一致
    """
    wb: Workbook = load_workbook(file_path)
    errors: list[ValidationError] = []
    target_sheets = sheets if sheets else [ws.title for ws in wb.worksheets]

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        has_header = _has_header_row(ws)

        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "f":
                    continue

                formula = safe_formula_str(cell.value)
                refs = extract_cell_references(formula)

                for ref in refs:
                    # 检查 1: 引用范围远超实际行数
                    if ref.max_row is not None and max_row > 0 and ref.max_row > max_row * 2:
                        errors.append(ValidationError(
                            severity="warning",
                            category="reference",
                            sheet=sheet_name,
                            cell=cell.coordinate,
                            code=ErrorCode.REF_OUT_OF_RANGE,
                            message=(
                                f"引用范围远超实际数据: {cell.coordinate} 引用行 {ref.max_row}"
                                f"，实际数据行 {max_row}"
                            ),
                            detail={
                                "formula": formula,
                                "ref_max_row": ref.max_row,
                                "data_max_row": max_row,
                            },
                        ))

                    # 检查 2: 是否包含表头行（第 1 行）
                    if has_header and ref.min_row == 1:
                        errors.append(ValidationError(
                            severity="warning",
                            category="reference",
                            sheet=sheet_name,
                            cell=cell.coordinate,
                            code=ErrorCode.REF_HEADER_INCLUDED,
                            message=f"公式包含表头行: {cell.coordinate}={formula}",
                            detail={
                                "formula": formula,
                                "ref_min_row": ref.min_row,
                            },
                        ))

                    # 检查 3: 聚合函数范围太小
                    if is_aggregate_function(formula) and ref.row_count <= 2:
                        errors.append(ValidationError(
                            severity="warning",
                            category="reference",
                            sheet=sheet_name,
                            cell=cell.coordinate,
                            code=ErrorCode.REF_INSUFFICIENT_RANGE,
                            message=(
                                f"聚合函数范围过小: {cell.coordinate}={formula}"
                                f"（范围仅 {ref.row_count} 单元格）"
                            ),
                            detail={
                                "formula": formula,
                                "row_count": ref.row_count,
                            },
                        ))

    # 检查 4: 同列公式模式一致性
    errors.extend(_check_formula_consistency(wb, target_sheets))

    wb.close()
    return errors
