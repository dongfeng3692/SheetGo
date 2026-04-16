"""公式错误检测 — 检测所有公式单元格的错误值"""

from __future__ import annotations

import re
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from .result import ErrorCode, ValidationError, map_error_code, safe_formula_str

if TYPE_CHECKING:
    from openpyxl import Workbook

# 隐式数组公式模式：在 LibreOffice 中可以直接计算，但在 MS Excel 中需要 CSE
# 这些函数在 Excel 中如果不按 Ctrl+Shift+Enter 会产生隐式数组行为
_IMPLICIT_ARRAY_PATTERNS: list[re.Pattern[str]] = [
    re.compile(r'\bMATCH\s*\(', re.IGNORECASE),
    re.compile(r'\bIF\s*\([^)]*,\s*[^,]+:[^,]+', re.IGNORECASE),  # IF 返回范围
    re.compile(r'\bINDEX\s*\([^)]*,\s*0\s*[,)]', re.IGNORECASE),  # INDEX 整列/行
    re.compile(r'\bSMALL\s*\(', re.IGNORECASE),
    re.compile(r'\bLARGE\s*\(', re.IGNORECASE),
    re.compile(r'\bFREQUENCY\s*\(', re.IGNORECASE),
    re.compile(r'\bTRANSPOSE\s*\(', re.IGNORECASE),
    re.compile(r'\bMMULT\s*\(', re.IGNORECASE),
    re.compile(r'\bMINVERSE\s*\(', re.IGNORECASE),
]

# 用于替换隐式数组为 SUMPRODUCT 包装的建议
_ARRAY_SAFE_ALTERNATIVES: dict[str, str] = {
    "MATCH": "使用 SUMPRODUCT 替代以避免 CSE",
    "IF": "使用 IF+INDEX/MATCH 替代数组 IF",
    "INDEX": "使用 INDEX 的单值形式",
    "SMALL": "使用 AGGREGATE(15,...) 替代",
    "LARGE": "使用 AGGREGATE(14,...) 替代",
    "FREQUENCY": "使用 COUNTIFS 替代",
    "TRANSPOSE": "使用 INDEX 逐个引用",
    "MMULT": "无直接替代，需 CSE 确认",
    "MINVERSE": "无直接替代，需 CSE 确认",
}


def is_implicit_array_formula(formula: str) -> bool:
    """检测公式是否为隐式数组公式（LibreOffice 兼容但 Excel 需要 CSE）"""
    if not formula or not formula.startswith("="):
        return False
    for pattern in _IMPLICIT_ARRAY_PATTERNS:
        if pattern.search(formula):
            return True
    return False


def _get_array_suggestion(formula: str) -> str:
    """获取隐式数组公式的替代建议"""
    for func, suggestion in _ARRAY_SAFE_ALTERNATIVES.items():
        if re.search(rf'\b{func}\s*\(', formula, re.IGNORECASE):
            return suggestion
    return "需要 Ctrl+Shift+Enter (CSE) 确认"


def check_formulas(
    file_path: str,
    sheets: list[str] | None = None,
) -> list[ValidationError]:
    """
    检测所有公式单元格的错误值。

    使用 openpyxl (data_only=True) 读取缓存的计算结果，
    检测 7 种标准 Excel 错误 + 隐式数组公式。
    """
    wb: Workbook = load_workbook(file_path, data_only=True)
    wb_formulas: Workbook = load_workbook(file_path, data_only=False)
    errors: list[ValidationError] = []

    target_sheets = sheets if sheets else [ws.title for ws in wb.worksheets]

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws_formulas = wb_formulas[sheet_name]

        for row in ws.iter_rows():
            for cell in row:
                # 检查错误类型单元格
                if cell.data_type == "e":
                    error_value = cell.value or ""
                    formula_cell = ws_formulas[cell.coordinate]
                    formula_text = safe_formula_str(formula_cell.value) if formula_cell.data_type == "f" else ""

                    errors.append(ValidationError(
                        severity="error",
                        category="formula",
                        sheet=sheet_name,
                        cell=cell.coordinate,
                        code=map_error_code(str(error_value)),
                        message=f"公式错误 {error_value} 在 {sheet_name}!{cell.coordinate}",
                        detail={
                            "error_value": str(error_value),
                            "formula": formula_text,
                        },
                    ))

                # 检查隐式数组公式（在公式工作簿中检查）
                formula_cell = ws_formulas[cell.coordinate]
                if formula_cell.data_type == "f":
                    formula = safe_formula_str(formula_cell.value)
                    if is_implicit_array_formula(formula):
                        errors.append(ValidationError(
                            severity="warning",
                            category="formula",
                            sheet=sheet_name,
                            cell=cell.coordinate,
                            code=ErrorCode.FORMULA_IMPLICIT_ARRAY,
                            message=f"隐式数组公式，需 CSE 确认: {formula}",
                            detail={
                                "formula": formula,
                                "suggestion": _get_array_suggestion(formula),
                            },
                        ))

    wb.close()
    wb_formulas.close()
    return errors
