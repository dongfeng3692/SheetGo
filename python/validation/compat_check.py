"""函数兼容性检查 — 检测不兼容函数"""

from __future__ import annotations

import re

from openpyxl import load_workbook

from .result import ErrorCode, ValidationError, safe_formula_str

# Excel 2019 及更早版本不支持的函数
FORBIDDEN_FUNCTIONS: dict[str, dict[str, str]] = {
    "FILTER":      {"alt": "SUMIF/COUNTIF + AutoFilter", "version": "2021+"},
    "UNIQUE":      {"alt": "Remove Duplicates + COUNTIF", "version": "2021+"},
    "SORT":        {"alt": "Data → Sort", "version": "2021+"},
    "SORTBY":      {"alt": "Data → Sort", "version": "2021+"},
    "XLOOKUP":     {"alt": "INDEX + MATCH", "version": "2021+"},
    "XMATCH":      {"alt": "MATCH", "version": "2021+"},
    "SEQUENCE":    {"alt": "ROW() or manual fill", "version": "2021+"},
    "LET":         {"alt": "Helper cells", "version": "2021+"},
    "LAMBDA":      {"alt": "Named ranges or VBA", "version": "2021+"},
    "RANDARRAY":   {"alt": "RAND() with fill", "version": "2021+"},
    "ARRAYFORMULA": {"alt": "CSE (Ctrl+Shift+Enter)", "note": "Google Sheets only"},
    "QUERY":       {"alt": "SUMIF/COUNTIF/PivotTable", "note": "Google Sheets only"},
}

# 所有禁用函数名集合，用于快速查找
_FORBIDDEN_NAMES = frozenset(FORBIDDEN_FUNCTIONS.keys())

# 函数名匹配正则（从公式中提取所有函数调用）
_FUNC_NAME_RE = re.compile(r'\b([A-Z]+)\s*\(', re.IGNORECASE)


def extract_all_function_names(formula: str) -> list[str]:
    """从公式中提取所有函数名（包括嵌套的）"""
    if not formula:
        return []
    return [m.group(1).upper() for m in _FUNC_NAME_RE.finditer(formula)]


def check_compatibility(
    file_path: str,
    sheets: list[str] | None = None,
) -> list[ValidationError]:
    """检测不兼容函数"""
    wb = load_workbook(file_path)
    errors: list[ValidationError] = []
    target_sheets = sheets if sheets else [ws.title for ws in wb.worksheets]

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "f":
                    continue

                formula = safe_formula_str(cell.value)
                func_names = extract_all_function_names(formula)

                for func_name in func_names:
                    if func_name in _FORBIDDEN_NAMES:
                        info = FORBIDDEN_FUNCTIONS[func_name]
                        version_info = info.get("version", "")
                        note = info.get("note", "")
                        msg = f"不兼容函数 {func_name}"
                        if version_info:
                            msg += f"（需要 Excel {version_info}）"
                        if note:
                            msg += f"（{note}）"

                        errors.append(ValidationError(
                            severity="error",
                            category="compat",
                            sheet=sheet_name,
                            cell=cell.coordinate,
                            code=ErrorCode.COMPAT_FORBIDDEN_FUNCTION,
                            message=msg,
                            detail={
                                "function": func_name,
                                "alternative": info["alt"],
                                "formula": formula,
                                "version": version_info,
                            },
                        ))

    wb.close()
    return errors
