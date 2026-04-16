"""sheet_info — 获取工作簿元信息"""

from __future__ import annotations

from typing import Any

from openpyxl import load_workbook

from .base import BaseTool


class SheetInfoTool(BaseTool):

    @property
    def name(self) -> str:
        return "sheet_info"

    @property
    def description(self) -> str:
        return (
            "获取 Excel 工作簿的元信息：所有工作表名称、行列数、"
            "合并单元格范围。用于了解文件整体结构。"
        )

    @property
    def parameters_schema(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Excel 文件路径",
                },
            },
            "required": ["file_path"],
        }

    async def execute(self, file_path: str, **kwargs) -> dict:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheets = []
        for ws in wb.worksheets:
            rows = ws.max_row or 0
            cols = ws.max_column or 0
            # ReadOnlyWorksheet 没有 dimensions/merged_cells，需要兼容
            try:
                dims = str(ws.dimensions) if ws.dimensions else ""
            except AttributeError:
                dims = f"A1:{_col_letter(cols)}{rows}" if rows and cols else ""
            try:
                merged = [str(m) for m in ws.merged_cells.ranges]
            except AttributeError:
                merged = []
            sheets.append({
                "name": ws.title,
                "rows": rows,
                "columns": cols,
                "dimensions": dims,
                "merged_cells": merged,
            })
        wb.close()
        return {"sheets": sheets}


def _col_letter(n: int) -> str:
    """1 → A, 26 → Z, 27 → AA"""
    result = ""
    while n > 0:
        n -= 1
        result = chr(65 + n % 26) + result
        n //= 26
    return result
