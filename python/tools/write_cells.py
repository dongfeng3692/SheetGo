"""写入类工具 + QueryData + ReadFormulas — 对接模块 4 Excel Engine"""

from __future__ import annotations

from typing import Any

from .base import BaseTool


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_DATE_FORMATS = [
    "%Y-%m-%d",
    "%Y-%m-%d %H:%M:%S",
    "%m/%d/%Y",
    "%d/%m/%Y",
    "%Y/%m/%d",
]


def _coerce_value(val: Any) -> Any:
    """Auto-detect date strings and convert to datetime for proper Excel storage."""
    if not isinstance(val, str):
        return val
    for fmt in _DATE_FORMATS:
        try:
            import datetime
            return datetime.datetime.strptime(val, fmt)
        except (ValueError, TypeError):
            continue
    return val

def _edit_result_to_dict(result: Any) -> dict:
    """将 EditResult 序列化为 JSON 友好的 dict。"""
    return {
        "success": result.success,
        "affected_cells": result.affected_cells,
        "affected_formulas": result.affected_formulas,
        "warnings": result.warnings,
    }


def _get_merged_non_topleft(file_path: str, sheet_name: str) -> set[str]:
    """Return set of cell refs (e.g. {'B1', 'C1'}) that are inside merged ranges but NOT the top-left cell."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    cells: set[str] = set()
    try:
        # read_only=True doesn't support merged_cells, use data_only=False
        wb = load_workbook(file_path, read_only=False)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return cells
        ws = wb[sheet_name]
        try:
            ranges = ws.merged_cells.ranges
        except AttributeError:
            wb.close()
            return cells
        for mr in ranges:
            for row in range(mr.min_row, mr.max_row + 1):
                for col in range(mr.min_col, mr.max_col + 1):
                    if row == mr.min_row and col == mr.min_col:
                        continue  # skip top-left
                    cells.add(f"{get_column_letter(col)}{row}")
        wb.close()
    except Exception:
        pass
    return cells


def _parse_range_str(range_str: str, sheet: str) -> Any:
    """将 'A1:D10' 解析为 CellRange。"""
    from excel.models import CellRange, parse_cell_ref

    parts = range_str.split(":")
    start_col, start_row = parse_cell_ref(parts[0])
    if len(parts) > 1:
        end_col, end_row = parse_cell_ref(parts[1])
    else:
        end_col, end_row = start_col, start_row
    return CellRange(
        sheet=sheet,
        start_col=start_col,
        start_row=start_row,
        end_col=end_col,
        end_row=end_row,
    )


# ---------------------------------------------------------------------------
# 写入类工具
# ---------------------------------------------------------------------------


class WriteCellsTool(BaseTool):
    @property
    def name(self) -> str:
        return "write_cells"

    @property
    def description(self) -> str:
        return "写入数据到指定单元格范围。保留原有格式和公式。"

    @property
    def parameters_schema(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Excel 文件路径",
                },
                "sheet": {"type": "string", "description": "工作表名"},
                "range": {"type": "string", "description": "Excel 范围，如 'A1:D10'"},
                "values": {
                    "type": "array",
                    "items": {"type": "array"},
                    "description": "二维数组，每行对应一行数据",
                },
            },
            "required": ["file_path", "sheet", "range", "values"],
        }

    @property
    def safe_level(self) -> str:
        return "write"

    async def execute(
        self,
        file_path: str,
        sheet: str,
        range: str,
        values: list[list[Any]],
        **kwargs,
    ) -> Any:
        from excel.models import CellEdit, col_number, col_letter, parse_cell_ref
        from excel.writer import ExcelWriter

        start_col, start_row = parse_cell_ref(range.split(":")[0])

        # Build merge set to protect merged cells
        merged_non_topleft = _get_merged_non_topleft(file_path, sheet)

        edits: list[CellEdit] = []
        warnings: list[str] = []
        for row_idx, row_values in enumerate(values):
            for col_idx, val in enumerate(row_values):
                if val is None:
                    continue
                cell_ref = f"{col_letter(col_number(start_col) + col_idx)}{start_row + row_idx}"
                if cell_ref in merged_non_topleft:
                    warnings.append(f"跳过合并单元格 {cell_ref} (值: {val!r})")
                    continue
                val = _coerce_value(val)
                edits.append(CellEdit(sheet=sheet, cell=cell_ref, value=val))

        writer = ExcelWriter()
        result = writer.write_cells(file_path, edits)
        result_dict = _edit_result_to_dict(result)
        if warnings:
            result_dict["warnings"] = result_dict.get("warnings", []) + warnings
        return result_dict


class AddFormulaTool(BaseTool):
    @property
    def name(self) -> str:
        return "add_formula"

    @property
    def description(self) -> str:
        return "在指定单元格添加公式。自动检测兼容性。"

    @property
    def parameters_schema(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Excel 文件路径",
                },
                "sheet": {"type": "string", "description": "工作表名"},
                "cell": {"type": "string", "description": "单元格位置，如 'C2'"},
                "formula": {"type": "string", "description": "公式，如 '=SUM(A2:A10)'"},
            },
            "required": ["file_path", "sheet", "cell", "formula"],
        }

    @property
    def safe_level(self) -> str:
        return "write"

    async def execute(
        self,
        file_path: str,
        sheet: str,
        cell: str,
        formula: str,
        **kwargs,
    ) -> Any:
        from excel.writer import ExcelWriter

        writer = ExcelWriter()
        result = writer.add_formula(file_path, sheet, cell, formula)
        return _edit_result_to_dict(result)


class AddColumnTool(BaseTool):
    @property
    def name(self) -> str:
        return "add_column"

    @property
    def description(self) -> str:
        return "在指定位置添加新列，可设置表头和数据或公式。"

    @property
    def parameters_schema(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Excel 文件路径",
                },
                "sheet": {"type": "string", "description": "工作表名"},
                "col_letter": {"type": "string", "description": "列字母，如 'D'"},
                "header": {"type": "string", "description": "列标题"},
                "data": {"type": "array", "description": "数据值列表"},
                "formula": {"type": "string", "description": "公式模板，用 {row} 表示当前行号"},
            },
            "required": ["file_path", "sheet", "col_letter", "header"],
        }

    @property
    def safe_level(self) -> str:
        return "write"

    async def execute(
        self,
        file_path: str,
        sheet: str,
        col_letter: str,
        header: str | None = None,
        data: list[Any] | None = None,
        formula: str | None = None,
        **kwargs,
    ) -> Any:
        from excel.writer import ExcelWriter

        writer = ExcelWriter()
        result = writer.add_column(
            file_path, sheet, col=col_letter, header=header, data=data, formula=formula,
        )
        return _edit_result_to_dict(result)


class InsertRowTool(BaseTool):
    @property
    def name(self) -> str:
        return "insert_row"

    @property
    def description(self) -> str:
        return "在指定位置插入行，自动调整公式引用。"

    @property
    def parameters_schema(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Excel 文件路径",
                },
                "sheet": {"type": "string", "description": "工作表名"},
                "at_row": {"type": "integer", "description": "插入位置（行号）"},
                "values": {
                    "type": "object",
                    "description": "列名到值的映射，如 {\"A\": 100, \"B\": \"text\"}",
                },
                "formula": {
                    "type": "object",
                    "description": "列名到公式的映射，如 {\"C\": \"=A{row}+B{row}\"}",
                },
            },
            "required": ["file_path", "sheet", "at_row"],
        }

    @property
    def safe_level(self) -> str:
        return "write"

    async def execute(
        self,
        file_path: str,
        sheet: str,
        at_row: int,
        values: dict[str, Any] | None = None,
        formula: dict[str, str] | None = None,
        **kwargs,
    ) -> Any:
        from excel.writer import ExcelWriter

        writer = ExcelWriter()
        result = writer.insert_row(file_path, sheet, at_row, values, formula)
        return _edit_result_to_dict(result)


class CreateChartTool(BaseTool):
    @property
    def name(self) -> str:
        return "create_chart"

    @property
    def description(self) -> str:
        return "创建图表并插入到工作表。支持柱状图、折线图、饼图等。"

    @property
    def parameters_schema(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Excel 文件路径",
                },
                "sheet": {"type": "string", "description": "数据源工作表"},
                "chart_type": {
                    "type": "string",
                    "enum": ["bar", "line", "pie", "area", "scatter"],
                    "description": "图表类型",
                },
                "data_range": {"type": "string", "description": "数据范围，如 'A1:D10'"},
                "target_cell": {"type": "string", "description": "图表插入位置"},
                "title": {"type": "string", "description": "图表标题"},
            },
            "required": ["file_path", "sheet", "chart_type", "data_range", "target_cell"],
        }

    @property
    def safe_level(self) -> str:
        return "write"

    async def execute(
        self,
        file_path: str,
        sheet: str,
        chart_type: str,
        data_range: str,
        target_cell: str,
        title: str | None = None,
        **kwargs,
    ) -> Any:
        from excel.chart_engine import ChartEngine
        from excel.models import ChartConfig

        source_range = _parse_range_str(data_range, sheet)
        config = ChartConfig(
            chart_type=chart_type,
            source_range=source_range,
            target_cell=target_cell,
            target_sheet=sheet,
            title=title or "",
        )
        result = ChartEngine.create_chart(file_path, config)
        return _edit_result_to_dict(result)


class ApplyStyleTool(BaseTool):
    @property
    def name(self) -> str:
        return "apply_style"

    @property
    def description(self) -> str:
        return "应用样式到指定范围（字体颜色、数字格式、边框等）。"

    @property
    def parameters_schema(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Excel 文件路径",
                },
                "sheet": {"type": "string", "description": "工作表名"},
                "range": {"type": "string", "description": "样式范围，如 'A1:D10'"},
                "style": {
                    "type": "object",
                    "properties": {
                        "role": {
                            "type": "string",
                            "enum": ["input", "formula", "cross_sheet", "header"],
                            "description": "样式角色",
                        },
                        "numfmt_type": {
                            "type": "string",
                            "enum": ["general", "currency", "percent", "integer"],
                            "description": "数字格式类型",
                        },
                    },
                    "required": ["role"],
                    "description": "样式配置",
                },
            },
            "required": ["file_path", "sheet", "range", "style"],
        }

    @property
    def safe_level(self) -> str:
        return "write"

    async def execute(
        self,
        file_path: str,
        sheet: str,
        range: str,
        style: dict,
        **kwargs,
    ) -> Any:
        from excel.models import parse_cell_ref
        from excel.style_engine import StyleEngine

        parts = range.split(":")
        start_col, start_row = parse_cell_ref(parts[0])
        end_col, end_row = parse_cell_ref(parts[1]) if len(parts) > 1 else (start_col, start_row)

        engine = StyleEngine()
        role = style.get("role", "input")
        numfmt_type = style.get("numfmt_type")

        result = engine.apply_financial_format(
            file_path, sheet, start_col, start_row, end_col, end_row,
            role=role, numfmt_type=numfmt_type,
        )
        return _edit_result_to_dict(result)


class ExportFileTool(BaseTool):
    @property
    def name(self) -> str:
        return "export_file"

    @property
    def description(self) -> str:
        return "导出文件为 xlsx 或 csv 格式。"

    @property
    def parameters_schema(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "源文件路径",
                },
                "output_path": {
                    "type": "string",
                    "description": "导出路径",
                },
                "format": {
                    "type": "string",
                    "enum": ["xlsx", "csv"],
                    "default": "xlsx",
                },
            },
            "required": ["file_path", "output_path"],
        }

    @property
    def safe_level(self) -> str:
        return "write"

    async def execute(
        self,
        file_path: str,
        output_path: str,
        format: str = "xlsx",
        **kwargs,
    ) -> Any:
        import shutil

        import pandas as pd

        if format == "xlsx":
            shutil.copy2(file_path, output_path)
            return {"success": True, "output_path": output_path}
        elif format == "csv":
            df = pd.read_excel(file_path, engine="openpyxl")
            df.to_csv(output_path, index=False)
            return {"success": True, "output_path": output_path}
        else:
            return {"success": False, "error": f"不支持的格式: {format}"}


# ---------------------------------------------------------------------------
# 查询类工具
# ---------------------------------------------------------------------------


class QueryDataTool(BaseTool):
    @property
    def name(self) -> str:
        return "query_data"

    @property
    def description(self) -> str:
        return (
            "用 SQL 查询 Excel 数据。传入 Excel 文件路径，自动加载所有工作表到内存 DuckDB。"
            "每个工作表名即表名，仅支持 SELECT 语句。"
            "列名使用字母 A, B, C, D...（与 read_sheet 一致），数值列自动转为数字类型。"
            "例如: SELECT A, SUM(E) FROM RANGES WHERE C LIKE 'BSDER%' GROUP BY A"
        )

    @property
    def parameters_schema(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Excel 文件路径",
                },
                "sql": {
                    "type": "string",
                    "description": "SQL 查询语句（仅支持 SELECT）",
                },
                "max_rows": {
                    "type": "integer",
                    "description": "返回最大行数（默认 100）",
                    "default": 100,
                },
            },
            "required": ["file_path", "sql"],
        }

    @property
    def safe_level(self) -> str:
        return "read"

    async def execute(self, file_path: str, sql: str, max_rows: int = 100, **kwargs) -> Any:
        import duckdb
        import pandas as pd
        from excel.duckdb_query import DuckDBQuery

        is_valid, error = DuckDBQuery.validate_sql(sql)
        if not is_valid:
            return {"error": error}

        # Load all sheets into an in-memory DuckDB
        # Use column letters (A, B, C...) to match read_sheet output
        con = duckdb.connect(":memory:")
        try:
            xlsx = pd.ExcelFile(file_path, engine="openpyxl")
            sheet_schemas: dict[str, list[str]] = {}
            for sheet_name in xlsx.sheet_names:
                df = pd.read_excel(xlsx, sheet_name=sheet_name, header=None)
                from tools.read_sheet import _col_letter
                df.columns = [_col_letter(i) for i in range(len(df.columns))]

                # Auto-cast: try to convert each column to numeric
                for col in df.columns:
                    converted = pd.to_numeric(df[col], errors="coerce")
                    # If >50% of non-null values converted successfully, use numeric
                    non_null = df[col].notna()
                    if non_null.any():
                        converted_valid = converted.notna() & non_null
                        if converted_valid.sum() / non_null.sum() > 0.5:
                            df[col] = converted

                con.register(sheet_name, df)

                # Build schema description for error messages
                col_types = []
                for col in df.columns:
                    dtype = str(df[col].dtype)
                    col_types.append(f"{col} ({dtype})")
                sheet_schemas[sheet_name] = col_types

            result = con.execute(sql).fetchdf()
        except duckdb.Error as e:
            # Enhanced error: include available columns and types
            schema_hint = ""
            if sheet_schemas:
                lines = []
                for sname, cols in sheet_schemas.items():
                    lines.append(f"  {sname}: {', '.join(cols)}")
                schema_hint = "\n\nAvailable columns:\n" + "\n".join(lines)
            return {"error": f"SQL execution error: {e}{schema_hint}"}
        finally:
            con.close()

        # Build column type info for the result
        col_type_info = []
        for col in result.columns:
            dtype = str(result[col].dtype)
            col_type_info.append(f"{col}:{dtype}")

        return {
            "columns": result.columns.tolist(),
            "types": col_type_info,
            "data": result.head(max_rows).values.tolist(),
            "total_rows": len(result),
            "truncated": len(result) > max_rows,
        }


# ---------------------------------------------------------------------------
# 读取类工具（已实现）
# ---------------------------------------------------------------------------


class ReadFormulasTool(BaseTool):
    @property
    def name(self) -> str:
        return "read_formulas"

    @property
    def description(self) -> str:
        return "读取工作表中所有公式及其位置。"

    @property
    def parameters_schema(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Excel 文件路径",
                },
                "sheet": {
                    "type": "string",
                    "description": "工作表名（可选，默认全部）",
                },
            },
            "required": ["file_path"],
        }

    async def execute(self, file_path: str, sheet: str | None = None, **kwargs) -> Any:
        from openpyxl import load_workbook

        wb = load_workbook(file_path)
        formulas = []
        target_sheets = [sheet] if sheet else [ws.title for ws in wb.worksheets]
        for s in target_sheets:
            if s not in wb.sheetnames:
                continue
            ws = wb[s]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        formulas.append({
                            "sheet": s,
                            "cell": cell.coordinate,
                            "formula": cell.value,
                        })
        wb.close()
        return {"formulas": formulas}
