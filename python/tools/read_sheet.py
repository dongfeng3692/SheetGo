"""read_sheet — 读取工作表数据（含 Excel 行号、列字母、合并单元格标注）"""

from __future__ import annotations

import math
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from .base import BaseTool


def _col_letter(n: int) -> str:
    """0-based column index → column letter (0→A, 25→Z, 26→AA)."""
    result = ""
    n += 1  # convert to 1-based
    while n > 0:
        n -= 1
        result = chr(65 + n % 26) + result
        n //= 26
    return result


def _col_number(letter: str) -> int:
    """Column letter → 1-based column number (A→1, Z→26, AA→27)."""
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _build_merge_map(file_path: str, sheet_name: str) -> dict[tuple[int, int], str]:
    """Build a map of (row, col_1based) → merge_range for all cells inside merged ranges.

    Only non-top-left cells get an entry. Value is the merge range string (e.g. "A1:C1").
    """
    # read_only=True doesn't support merged_cells, use data_only=False
    wb = load_workbook(file_path)
    try:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]
        merge_map: dict[tuple[int, int], str] = {}
        try:
            ranges = ws.merged_cells.ranges
        except AttributeError:
            return {}

        for mr in ranges:
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    if r != mr.min_row or c != mr.min_col:
                        merge_map[(r, c)] = str(mr)
        return merge_map
    finally:
        wb.close()


class ReadSheetTool(BaseTool):

    @property
    def name(self) -> str:
        return "read_sheet"

    @property
    def description(self) -> str:
        return (
            "读取 Excel 工作表的数据，返回每行的 Excel 行号和列字母映射。"
            "合并单元格只保留左上角的值，其余位置标注为 merged。"
            "默认返回前 100 行，可调整 max_rows。"
        )

    @property
    def parameters_schema(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Excel 文件路径",
                },
                "sheet": {
                    "type": "string",
                    "description": "工作表名（默认第一个）",
                },
                "max_rows": {
                    "type": "integer",
                    "description": "返回最大行数（默认 100）",
                    "default": 100,
                },
            },
            "required": ["file_path"],
        }

    async def execute(
        self,
        file_path: str,
        sheet: str | None = None,
        max_rows: int = 100,
        **kwargs,
    ) -> dict:
        # Resolve sheet name first
        sheet_name = sheet or ""
        if not sheet_name:
            wb_tmp = load_workbook(file_path, read_only=True)
            sheet_name = wb_tmp.sheetnames[0]
            wb_tmp.close()

        # Build merge map for annotations
        merge_map = _build_merge_map(file_path, sheet_name)

        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            engine="openpyxl",
            header=None,
            nrows=max_rows,
        )
        col_letters = [_col_letter(i) for i in range(len(df.columns))]

        # Collect merge ranges that appear in the returned data
        seen_merges: list[str] = []

        rows: list[dict] = []
        for idx in range(len(df)):
            row_num = idx + 1  # Excel rows are 1-based
            cells: dict[str, Any] = {}
            has_value = False
            last_non_null = -1  # rightmost column index with a real value
            raw: list[Any] = []
            for col_idx, val in enumerate(df.iloc[idx]):
                v = val
                col_1 = col_idx + 1

                # Check if this cell is inside a merged range (but not top-left)
                merge_ref = merge_map.get((row_num, col_1))
                if merge_ref is not None:
                    v = f"(merged: {merge_ref})"
                    if merge_ref not in seen_merges:
                        seen_merges.append(merge_ref)
                else:
                    # Convert NaN/None to None
                    if isinstance(v, float) and math.isnan(v):
                        v = None
                    elif pd.isna(v):
                        v = None

                if v is not None:
                    has_value = True
                    last_non_null = col_idx
                raw.append(v)

            # Trim trailing None cells
            cells: dict[str, Any] = {}
            for col_idx in range(last_non_null + 1):
                if raw[col_idx] is not None:
                    cells[col_letters[col_idx]] = raw[col_idx]
            if has_value:
                rows.append({"row": row_num, "cells": cells})

        return {
            "sheet": sheet_name,
            "columns": col_letters,
            "rows": rows,
            "merged_ranges": seen_merges,
            "total_rows": len(df),
            "truncated": False,
        }
