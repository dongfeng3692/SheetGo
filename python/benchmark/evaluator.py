"""Evaluator — SpreadsheetBench 评测逻辑（自包含，不依赖 SpreadsheetBench 仓库）

适配自 SpreadsheetBench evaluation.py，核心逻辑:
- openpyxl data_only=True 读取缓存值
- 在 answer_position 指定范围内逐格比对
- 数值保留 2 位小数，空字符串等价 None，类型须匹配
"""

from __future__ import annotations

import datetime
import json
import math
import os
from typing import Any

import openpyxl
from openpyxl.utils import range_boundaries


# ============================================================================
# 单元格值比较
# ============================================================================


def _transform_value(value: Any) -> Any:
    """标准化单元格值以便比较"""
    if value is None or value == "":
        return None

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return round(float(value), 2)

    if isinstance(value, datetime.datetime):
        # 转为 Excel 序列日期数
        delta = value - datetime.datetime(1899, 12, 30)
        return round(delta.total_seconds() / 86400, 0)

    if isinstance(value, datetime.time):
        # 截断到 HH:MM
        return value.strftime("%H:%M")

    if isinstance(value, str):
        # 尝试转为数值
        try:
            return round(float(value), 2)
        except (ValueError, TypeError):
            return value

    return value


def compare_cell_value(v1: Any, v2: Any) -> bool:
    """比较两个单元格值是否相等"""
    t1 = _transform_value(v1)
    t2 = _transform_value(v2)

    # None 和空字符串等价
    if t1 is None and t2 is None:
        return True
    if t1 is None or t2 is None:
        return False

    # 类型必须匹配
    if type(t1) != type(t2):
        return False

    return t1 == t2


# ============================================================================
# answer_position 解析
# ============================================================================


def _parse_answer_position(answer_position: str, default_sheet: str) -> list[tuple[str, str]]:
    """解析 answer_position 为 [(sheet_name, cell_range), ...]

    支持格式:
    - "A1"
    - "A1:B10"
    - "Sheet2!A1:C5"
    - "Sheet1!A1:B2,Sheet2!C3:D4"
    """
    parts = [p.strip() for p in answer_position.split(",")]
    result = []
    for part in parts:
        if "!" in part:
            sheet, cell_range = part.split("!", 1)
            sheet = sheet.strip().strip("'\"")
            cell_range = cell_range.strip()
        else:
            sheet = default_sheet
            cell_range = part.strip()
        result.append((sheet, cell_range))
    return result


def _expand_cell_range(cell_range: str) -> list[str]:
    """将范围展开为单元格列表，如 'A1:C3' → ['A1','B1','C1','A2',...]"""
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except (ValueError, TypeError):
        return [cell_range]

    cells = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            from openpyxl.utils import get_column_letter
            cells.append(f"{get_column_letter(col)}{row}")
    return cells


# ============================================================================
# 工作簿比较
# ============================================================================


def compare_workbooks(
    gt_path: str,
    output_path: str,
    answer_position: str,
    answer_sheet: str | None = None,
) -> tuple[bool, str]:
    """比较输出文件与 ground truth

    Returns:
        (passed, message)
    """
    if not os.path.exists(output_path):
        return False, "输出文件不存在"
    if not os.path.exists(gt_path):
        return False, "ground truth 文件不存在"

    try:
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    except Exception as e:
        return False, f"无法读取 ground truth: {e}"

    # Output file: NOT data_only — so we can evaluate formulas against cached values.
    # We compute formula values ourselves.
    try:
        out_wb = openpyxl.load_workbook(output_path, data_only=False)
    except Exception as e:
        gt_wb.close()
        return False, f"无法读取输出文件: {e}"

    # Build a simple value cache from output workbook for formula evaluation
    out_cache = _build_value_cache(out_wb)

    # 优先使用 answer_sheet，否则用 golden 的第一个 sheet
    default_sheet = answer_sheet or (gt_wb.sheetnames[0] if gt_wb.sheetnames else "Sheet1")

    ranges = _parse_answer_position(answer_position, default_sheet)

    for sheet_name, cell_range in ranges:
        gt_ws = gt_wb[sheet_name] if sheet_name in gt_wb.sheetnames else None
        out_ws = out_wb[sheet_name] if sheet_name in out_wb.sheetnames else None

        if gt_ws is None:
            continue
        if out_ws is None:
            gt_wb.close()
            out_wb.close()
            return False, f"工作表 '{sheet_name}' 不存在于输出文件中"

        cells = _expand_cell_range(cell_range)
        for cell_ref in cells:
            gt_val = gt_ws[cell_ref].value

            out_cell = out_ws[cell_ref]
            out_val = _resolve_cell_value(out_cell, out_cache, sheet_name)

            if not compare_cell_value(gt_val, out_val):
                gt_wb.close()
                out_wb.close()
                return False, (
                    f"单元格 {sheet_name}!{cell_ref} 不匹配: "
                    f"期望={gt_val!r}, 实际={out_val!r}"
                )

    gt_wb.close()
    out_wb.close()
    return True, "通过"


def _build_value_cache(wb: openpyxl.Workbook) -> dict[str, dict[str, Any]]:
    """Build a cache of all static (non-formula) cell values per sheet.

    Returns: {sheet_name: {cell_ref: value}}
    """
    cache: dict[str, dict[str, Any]] = {}
    for ws in wb.worksheets:
        sheet_cache: dict[str, Any] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and not (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    sheet_cache[cell.coordinate] = cell.value
        cache[ws.title] = sheet_cache
    return cache


def _resolve_cell_value(
    cell: Any,
    cache: dict[str, dict[str, Any]],
    sheet_name: str,
) -> Any:
    """Resolve a cell's effective value.

    If the cell is a formula, try to evaluate it using the value cache.
    Otherwise return the raw value.
    """
    val = cell.value
    if val is None:
        return None
    if isinstance(val, str) and val.startswith("="):
        return _eval_simple_formula(val, cache, sheet_name)
    return val


def _parse_ref(ref: str) -> tuple[str | None, str, int]:
    """Parse a cell reference like 'A1' or 'Sheet1!A1' into (sheet, col, row)."""
    sheet = None
    if "!" in ref:
        sheet_part, ref = ref.split("!", 1)
        sheet = sheet_part.strip().strip("'\"")
    col_str = ""
    row_str = ""
    for ch in ref:
        if ch.isalpha():
            col_str += ch
        else:
            row_str += ch
    return sheet, col_str, int(row_str) if row_str else 0


def _ref_to_idx(col_str: str, row: int) -> tuple[int, int]:
    """Convert column letter + row to 0-based (col, row)."""
    col = 0
    for ch in col_str.upper():
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return col - 1, row - 1


def _idx_to_col(col_idx: int) -> str:
    """0-based column index to letter."""
    result = ""
    col_idx += 1
    while col_idx > 0:
        col_idx -= 1
        result = chr(65 + col_idx % 26) + result
        col_idx //= 26
    return result


def _expand_range_ref(range_str: str, default_sheet: str | None) -> list[tuple[str | None, str]]:
    """Expand 'A1:B3' into [(sheet, 'A1'), (sheet, 'B1'), ...]."""
    parts = range_str.split(":")
    if len(parts) != 2:
        return [(default_sheet, range_str)]
    s1, c1, r1 = _parse_ref(parts[0])
    s2, c2, r2 = _parse_ref(parts[1])
    sheet = s1 or s2 or default_sheet
    col1, row1 = _ref_to_idx(c1, r1)
    col2, row2 = _ref_to_idx(c2, r2)
    result = []
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            result.append((sheet, f"{_idx_to_col(c)}{r + 1}"))
    return result


def _eval_simple_formula(
    formula: str,
    cache: dict[str, dict[str, Any]],
    current_sheet: str,
) -> Any:
    """Evaluate simple formulas (SUM, COUNT, AVERAGE, MIN, MAX) using cached values.

    Falls back to None for unsupported formulas.
    """
    import re

    expr = formula.lstrip("=")

    # Match function(range) pattern
    m = re.match(r"(\w+)\((.+)\)$", expr, re.IGNORECASE)
    if not m:
        return None

    func_name = m.group(1).upper()
    range_arg = m.group(2).strip()

    # Resolve values from cache
    refs = _expand_range_ref(range_arg, current_sheet)
    values: list[float] = []
    for sheet, cell_ref in refs:
        s = sheet or current_sheet
        sheet_cache = cache.get(s, {})
        v = sheet_cache.get(cell_ref)
        if v is not None:
            if isinstance(v, bool):
                continue
            if isinstance(v, (int, float)):
                values.append(float(v))
            elif isinstance(v, datetime.datetime):
                # Convert datetime to serial for SUM
                delta = v - datetime.datetime(1899, 12, 30)
                values.append(delta.total_seconds() / 86400)

    if not values:
        return 0 if func_name in ("SUM", "COUNT", "AVERAGE", "MIN", "MAX") else None

    if func_name == "SUM":
        return sum(values)
    elif func_name == "COUNT":
        return len(values)
    elif func_name == "AVERAGE":
        return sum(values) / len(values)
    elif func_name == "MIN":
        return min(values)
    elif func_name == "MAX":
        return max(values)

    return None


# ============================================================================
# 批量评测
# ============================================================================


def evaluate_task(
    task: dict,
    dataset_dir: str,
    output_dir: str,
    num_test_cases: int = 1,
) -> dict:
    """评测单个任务的所有 test case

    Returns:
        {
            "id": str,
            "instruction_type": str,
            "test_case_results": [1, 1, 0],   # 1=通过, 0=失败
            "soft_restriction": float,         # 通过率
            "hard_restriction": int,           # 全部通过=1
            "messages": [str, ...],            # 每个用例的评测消息
        }
    """
    task_id = str(task["id"])
    answer_position = task["answer_position"]
    answer_sheet = task.get("answer_sheet")
    spreadsheet_dir = os.path.join(dataset_dir, task["spreadsheet_path"])

    results = []
    messages = []

    for tc_idx in range(num_test_cases):
        tc_num = tc_idx + 1
        gt_path = os.path.join(spreadsheet_dir, f"{tc_num}_{task_id}_golden.xlsx")
        out_path = os.path.join(output_dir, f"{tc_num}_{task_id}_output.xlsx")

        if not os.path.exists(gt_path):
            results.append(0)
            messages.append(f"Test case {tc_num}: ground truth 不存在")
            continue

        passed, msg = compare_workbooks(gt_path, out_path, answer_position, answer_sheet)
        results.append(1 if passed else 0)
        messages.append(f"Test case {tc_num}: {msg}")

    total = len(results)
    passed = sum(results)
    soft = round(passed / total, 4) if total > 0 else 0.0
    hard = 1 if passed == total and total > 0 else 0

    return {
        "id": task_id,
        "instruction_type": task.get("instruction_type", "unknown"),
        "instruction": task.get("instruction", ""),
        "test_case_results": results,
        "soft_restriction": soft,
        "hard_restriction": hard,
        "messages": messages,
    }


def evaluate_dataset(
    dataset_dir: str,
    output_dir: str,
    tasks: list[dict],
) -> list[dict]:
    """批量评测整个数据集"""
    results = []
    for task in tasks:
        result = evaluate_task(task, dataset_dir, output_dir)
        results.append(result)
    return results


# ============================================================================
# 报告
# ============================================================================


def print_report(results: list[dict], model: str, dataset: str, max_tasks: int) -> None:
    """打印评测汇总报告"""
    total = len(results)
    if total == 0:
        print("没有评测结果")
        return

    soft_pass = sum(1 for r in results if r["soft_restriction"] > 0)
    hard_pass = sum(1 for r in results if r["hard_restriction"] == 1)

    soft_rate = soft_pass / total * 100
    hard_rate = hard_pass / total * 100

    # 按类型分组
    by_type: dict[str, list[dict]] = {}
    for r in results:
        t = r["instruction_type"]
        by_type.setdefault(t, []).append(r)

    sep = "=" * 60
    print(f"\n{sep}")
    print("SpreadsheetBench 评测报告")
    print(sep)
    print(f"模型:          {model}")
    print(f"数据集:        {dataset}")
    print(f"评测任务数:    {total}/{max_tasks or '全部'}")
    print()
    print("总体通过率:")
    print(f"  Soft (任意用例通过): {soft_rate:.1f}%  ({soft_pass}/{total})")
    print(f"  Hard (全部用例通过): {hard_rate:.1f}%  ({hard_pass}/{total})")
    print()

    if len(by_type) > 1:
        print("按任务类型:")
        for type_name, type_results in by_type.items():
            t_total = len(type_results)
            t_soft = sum(1 for r in type_results if r["soft_restriction"] > 0) / t_total * 100
            t_hard = sum(1 for r in type_results if r["hard_restriction"] == 1) / t_total * 100
            print(f"  {type_name}:  Soft {t_soft:.1f}% | Hard {t_hard:.1f}%")
        print()

    # 失败任务明细（最多显示 10 个）
    failed = [r for r in results if r["hard_restriction"] == 0]
    if failed:
        print(f"失败任务 ({len(failed)} 个):")
        for r in failed[:10]:
            tc_summary = "/".join(str(x) for x in r["test_case_results"])
            print(f"  #{r['id']} [{tc_summary}] {r['instruction'][:60]}...")
        if len(failed) > 10:
            print(f"  ... 还有 {len(failed) - 10} 个")

    print(sep)


def save_report(
    results: list[dict],
    model: str,
    dataset: str,
    output_path: str,
) -> None:
    """保存 JSON 报告"""
    total = len(results)
    soft_pass = sum(1 for r in results if r["soft_restriction"] > 0)
    hard_pass = sum(1 for r in results if r["hard_restriction"] == 1)

    report = {
        "model": model,
        "dataset": dataset,
        "total_tasks": total,
        "soft_pass_count": soft_pass,
        "hard_pass_count": hard_pass,
        "soft_pass_rate": round(soft_pass / total, 4) if total > 0 else 0,
        "hard_pass_rate": round(hard_pass / total, 4) if total > 0 else 0,
        "tasks": results,
    }

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
